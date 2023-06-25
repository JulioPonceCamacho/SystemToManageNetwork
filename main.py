from modulos import *

#Librerias funcionales de python
import os
import sys
from time import sleep
import subprocess as sb
from threading import Thread
from openpyxl import Workbook
from datetime import datetime as dt
from webbrowser import open as WebOpen
from scapy.all import sniff
from getpass import getuser
#Librerias de la interfaz Grafica
from PySide6.QtGui import QPixmap,QPainter,QIcon,QColor
from PySide6.QtWidgets import QApplication, QMainWindow, QHBoxLayout, QLabel,QWidget,QVBoxLayout, QGridLayout,QFrame,QPushButton, QTableWidgetItem
from PySide6.QtCharts import QBarSeries,QBarSet,QBarCategoryAxis,QChart,QLineSeries,QChartView,QPieSeries
from PySide6.QtCore import Qt,QEasingCurve,QPropertyAnimation,Signal,QFile,QTextStream
from modulos.UI_Interfaz import Ui_MainWindow
import socket
os.environ["QT_FONT_DPI"] = "96"

class Informacion_Red():
    INTERFAZ=""
    GATEWAY=""
    auxGate=""
    IPV4=""
    CAMBIOS=False
    SALIR=False
    SSID=""
    MASCARA=""
    Errores=[]
    VENTANA=""
    TEMA=""
    DETECCION=""
    TIEMPO_DETECCION=0
    TIEMPO_PAQ=0
    CONEXION=True
    def __init__(self) -> None:
        self.Iniciarconfiguracion()
        dis=rs.Obtener_Dispositivos(True)
        auxIns=[]
        for INTER in dis:
            if "WI-FI" in INTER[0].upper() or "WIFI" in INTER[0].upper() or "ETHER" in INTER[0].upper():
                auxIns.append(INTER[0])
            if INTER[3] != "--" and INTER[3] != "127.0.0.1" and INTER[5]!="--" and "VIRTUAL" not in INTER[0].upper():
                self.INTERFAZ=INTER[0]
                self.MASCARA=INTER[4]
                self.GATEWAY=INTER[5]
                self.IPV4=INTER[3]
                self.SSID=rs.obtener_SSID(self.INTERFAZ)
                if self.SSID=='NO TIENE':
                    try:
                        self.SSID=str(socket.getfqdn(self.GATEWAY))
                        if self.SSID==self.GATEWAY:
                            try:
                             self.SSID=str(rs.get_mac_details(rs.escanearARP_U(self.GATEWAY,self.INTERFAZ,2,False)[1]))
                            except:
                                self.SSID=self.GATEWAY
                    except:
                        try:
                             self.SSID=str(rs.get_mac_details(rs.escanearARP_U(self.GATEWAY,self.INTERFAZ,2,False)[1]))
                        except:
                            self.SSID=self.GATEWAY
                            self.auxGate=self.GATEWAY
                aux=db.consultarRED()
                encontrado=False
                for i in aux:
                    if i[1]==self.SSID:
                        encontrado=True
                if encontrado==False and self.SSID!='POR DEFECTO' and self.SSID != "SIN CONEXION" and self.SSID != "--" and self.SSID != "":
                    db.insertarRED(self.SSID,"Pendiente",0,0,1)
        if self.SSID=="SIN CONEXION":
            print("[!] No hay conexión con ninguna red.")
            self.IPV4="--"
            self.GATEWAY="--"
            self.CONEXION=False
            self.Errores.append([f"{dt.now().hour}:{dt.now().minute}","Sin red.","No hay conexión con ninguna red."])    
            e=False
            if len(auxIns)>0:
                for a in auxIns:
                    if "VIRTUAL" not in a.upper():
                        self.INTERFAZ=a
                        e=True
                        break
            if e==False:
                print("[!] No hay interfaces de red.")
                self.INTERFAZ="Wi-Fi"            
    def setIPV4(self,ipv4):
        self.IPV4=ipv4        
    def getIPV4(self):
        return self.IPV4
    def setGATEWAY(self,gate):
        self.GATEWAY=gate
    def getGATEWAY(self):
        return self.GATEWAY
    def setInterfaz(self,inter):
        self.INTERFAZ=inter
    def getInterfaz(self):
        return self.INTERFAZ
    def Iniciarconfiguracion(self):
        if os.path.exists('modulos/BasesDatos/configuracion.conf')==False:
            config=open("modulos/BasesDatos/configuracion.conf","w",encoding="utf-8")
            self.VENTANA="Propia del Software"
            config.write('Propia del Software\n')
            self.TEMA="Obscuro"
            config.write('Obscuro\n')
            self.DETECCION="Preciso (ARP y Ping)"
            config.write('Preciso (ARP y Ping)\n')
            self.TIEMPO_DETECCION=20.0
            config.write('20\n')
            self.TIEMPO_PAQ=0.5
            config.write('0.5')
            config.close()
        else:
            aux=open("modulos/BasesDatos/configuracion.conf","r",encoding="utf-8")
            contenido=aux.readlines()
            self.VENTANA=contenido[0].replace("\n","")
            self.TEMA=contenido[1].replace("\n","")
            self.DETECCION=contenido[2].replace("\n","")
            self.TIEMPO_DETECCION=float(contenido[3].replace("\n",""))
            self.TIEMPO_PAQ=float(contenido[4].replace("\n",""))
            aux.close()
    def verificarNpcap(self):
        if os.path.isdir('C:\\Program Files\\Npcap') or os.path.isdir('C:\\Program Files (x86)\\Npcap') or os.path.isdir('C:\\Archivos de programa\\Npcap') or os.path.isdir('C:\\Archivos de programa (x86)\\Npcap') or os.path.isdir('C:\\Program Files\\WinPcap') or os.path.isdir('C:\\Program Files (x86)\\WinPcap') or os.path.isdir('C:\\Archivos de programa\\WinPcap') or os.path.isdir('C:\\Archivos de programa (x86)\\WinPcap'):
            print("-- DEPENDENCIA: Winpcap esta instalado.")
        else:
            print("[i] DEPENDENCIA: Winpcap no esta instalado.")
            sb.run(('Dependencias/winpcap.exe'))
            os.execl(sys.executable, "main.py", *sys.argv)
    def verificarSpeed(self):
        if os.path.isdir(os.path.expanduser('~')+'\\AppData\\Roaming\\Ookla'):
            print("-- DEPENDENCIA: Terminos de Ookla aceptados.")
        else:
            print("[!] DEPENDENCIA: No se han aceptado los terminos de Ookla, corrigiendo.")
            print("\n")
            print("######################################################################")
            print("##   Para comenzar acepte los terminos de SpeedTest, coloque YES    ##")
            print("######################################################################")
            sb.run(["modulos/SpeedTest/speedtest.exe"])
        
class DispositivosClass(QFrame):
    INTER=None
    def __init__(self,INTER,tema,info):
        super(DispositivosClass,self).__init__()
        NOM=QLabel()
        self.INTER=INTER
        if INTER[3] != "--" or INTER[6] != "--":
            NOM.setText("<b>Nombre Interfaz:</b><br>->"+INTER[0]+"<br><b>ID:</b><br>->"+INTER[1]+"<br><b>MAC:</b><br>->"+INTER[2]+"<br>IPv4:</br><ul><li><b>IP:</b> "+INTER[3]+"</li><li><b>Sub-Máscara:</b>"+INTER[4]+"</li><li><b>Gateway:</b>"+INTER[5]+"</li></ul><b>IPV6:</b> "+INTER[6])
        elif INTER[3] != "--":
            NOM.setText("<b>Nombre Interfaz:</b><br>->"+INTER[0]+"<br><b>ID:</b><br>->"+INTER[1]+"<br><b>MAC:</b><br>->"+INTER[2]+"<br>IPv4:</br><ul><li><b>IP:</b> "+INTER[3]+"</li><li><b>Sub-Máscara:</b>"+INTER[4]+"</li><li><b>Gateway:</b>"+INTER[5]+"</li></ul>")
        elif INTER[6] != "--":
            NOM.setText("<b>Nombre Interfaz:</b><br>->"+INTER[0]+"<br><b>ID:</b><br>->"+INTER[1]+"<br><b>MAC:</b><br>->"+INTER[2]+"<br><b>IPV6:</b> "+INTER[6])
        else:
            NOM.setText("<b>Nombre Interfaz:</b><br>->"+INTER[0]+"<br><b>ID:</b><br>->"+INTER[1]+"<br><b>MAC:</b><br>->"+INTER[2])

        #BROADCAST=QLabel(INTER[7])
        Layout = QGridLayout()
        label = QLabel(self)
        if "Wi-Fi" in INTER[0] or "WI-FI" in INTER[0] or "wi-fi" in INTER[0] or "Wireless" in INTER[0]  or "WIRELESS" in INTER[0]  or "wireless" in INTER[0] or "wlan" in INTER[0] or "wlp" in INTER[0]:
            pixmap = QPixmap('Imagenes/INTER/WIFI.png').scaled(50,50)
        elif "Virtual" in INTER[0] or "VIRTUAL" in INTER[0] or "virtual" in INTER[0] or "VM" in INTER[0]:
            pixmap = QPixmap('Imagenes/INTER/VMW.png').scaled(50,50)
        elif "Local" in INTER[0] or "local" in INTER[0] or "LOCAL" in INTER[0]:
            pixmap = QPixmap('Imagenes/INTER/LAN.png').scaled(50,50)
        elif "BLUETOOTH" in INTER[0] or "Bluetooth" in INTER[0] or "bluetooth" in INTER[0] or "luetoo" in INTER[0]:
            pixmap = QPixmap('Imagenes/INTER/BLUE.png').scaled(50,50)
        else:
            pixmap = QPixmap('Imagenes/INTER/DESC.png').scaled(50,50)
        label.setPixmap(pixmap)
        Aux2=QWidget()
        Aux = QHBoxLayout()
        Aux2.setLayout(Aux)
        Aux.addWidget(QLabel(""))
        Aux.addWidget(label)
        Aux.addWidget(QLabel(""))
        Layout.addWidget(Aux2,0,0)
        Layout.addWidget(NOM,1,0)

        def SeleccionInter():
            info.setInterfaz(self.INTER[0])
            info.setIPV4(self.INTER[3])
            info.MASCARA=self.INTER[4]
            info.setGATEWAY(self.INTER[5])
            info.CAMBIOS=True
        if INTER[3] != "--" and INTER[3] != "127.0.0.1" and INTER[5]!="--":
            SELECT = QPushButton("Seleccionar Interfaz")
            if tema=="Obscuro":
                SELECT.setStyleSheet("background-color:#2c313c; border-radius:3px;")
            else:
                SELECT.setStyleSheet("background-color:#D8D8D8;border: 1px solid;border-color:#333333;")

            SELECT.clicked.connect(SeleccionInter)
            Layout.addWidget(SELECT,2,0)
        if tema=="Obscuro":
            self.setStyleSheet("QLabel{border-color:black;} QFrame{Background-color:#1f232a; Border-Radius:10px; }")
        else:
            self.setStyleSheet("QLabel{border-color:black;} QFrame{Background-color:#D8D8D8; Border-Radius:10px; }")
        self.setLayout(Layout)
class pendientesClass(QFrame):
    def __init__(self,INTER,tema,inf):
        super(pendientesClass,self).__init__()
        FECHA=INTER[2]
        DES=INTER[1]
        RED=INTER[6]
        Layout = QGridLayout()
        label = QLabel(self)
        pixmap = QPixmap('Imagenes/pendientes.png').scaled(50,50)
        label.setPixmap(pixmap)
        Aux2=QWidget()
        Aux = QHBoxLayout()
        Aux2.setLayout(Aux)
        Layout.addWidget(label,0,0)
        Layout.addWidget(QLabel(FECHA),0,1)
        Layout.addWidget(QLabel(RED),0,2)
        Layout.addWidget(QLabel(DES),0,3)
        if tema=="Obscuro":
            self.setStyleSheet("QLabel{border-color:black;} QFrame{Background-color:#1f232a; Border-Radius:10px; }")
        else:
            self.setStyleSheet("QLabel{border-color:black;} QFrame{Background-color:#D8D8D8; Border-Radius:10px; }")     
        self.setLayout(Layout)
class UsuariosClass(QFrame):
    def __init__(self,INTER,tema,inf):
        super(UsuariosClass,self).__init__()
        Layout = QGridLayout()
        label = QLabel(self)
        if INTER[0] == inf.GATEWAY:
            pixmap = QPixmap('Imagenes/DISP/ROUTER.png').scaled(50,50)
        elif INTER[0] != INTER[3]:
            pixmap = QPixmap('Imagenes/DISP/PC.png').scaled(50,50)
        else:
            pixmap = QPixmap('Imagenes/DISP/HOST.png').scaled(50,50)
        label.setPixmap(pixmap)
        NOM=QLabel()
        if INTER[0] == inf.IPV4:
            NOM.setText("<b>Usuario:</b> "+INTER[3]+" (Tu Dispositivo)<br><b>IP: </b>"+INTER[0]+"<br><b>MAC: </b>"+INTER[1]+"<br><b>Vendor: </b>"+INTER[2])
        else:
            NOM.setText("<b>Usuario: </b>"+INTER[3]+"<br><b>IP: </b>"+INTER[0]+"<br><b>MAC: </b>"+INTER[1]+"<br><b>Vendor: </b>"+INTER[2])
        NOM.wordWrap()
        Aux2=QWidget()
        Aux = QHBoxLayout()
        Aux2.setLayout(Aux)
        Aux.addWidget(QLabel(""))
        Aux.addWidget(label)
        Aux.addWidget(QLabel(""))
        Layout.addWidget(label,0,0)
        Layout.addWidget(NOM,1,0)
        if tema=="Obscuro":
            self.setStyleSheet("QLabel{border-color:black;} QFrame{Background-color:#1f232a; Border-Radius:10px; }")
        else:
            self.setStyleSheet("QLabel{border-color:black;} QFrame{Background-color:#D8D8D8; Border-Radius:10px; }")     
        self.setLayout(Layout)
class TraficoVar():
    VAR=["ID","T","P","L","S","D","E"]
    Cerrar=False
    Borrar=False
    velocidad=[]
class MainWindow(QMainWindow):
    signalUser = Signal(int)
    signalPing = Signal(int)
    signalTrafico = Signal(int)
    SignalV=Signal(int)
    tab=Signal(int)
    SignalVT=Signal(int)
    SignalVTL=Signal(int)
    SignalA=Signal(int)

    GLOBAL_STATE=False
    inf=Informacion_Red()
    TrafObj=TraficoVar()

    VEL=None
    def __init__(self):
        super(MainWindow, self).__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.GLOBAL_STATE=False
        ##QUITAR DECORACION 
        if self.inf.VENTANA=="Propia del Software":
            self.setWindowFlags(Qt.FramelessWindowHint)
        else:
            self.ui.SalirBtn.hide()
            self.ui.MinBtn.hide()
            self.ui.MaxBtn.hide()
        if self.inf.TEMA!="Obscuro":
            file = QFile("modulos/BasesDatos/claro.qss")
            file.open(QFile.ReadOnly | QFile.Text)
            stream = QTextStream(file)
            self.setStyleSheet(stream.readAll())
        self.inf.verificarNpcap()
        self.inf.verificarSpeed()
        ##############################################
        ## ACCION DE LOS BOTONES
        ##############################################
        self.ui.SalirBtn.clicked.connect(self.Cerrar)
        self.ui.MinBtn.clicked.connect(self.Minimizar)
        self.ui.MaxBtn.clicked.connect(self.Maximizar_Restaurar)

        self.ui.Menu.clicked.connect(lambda: self.Menu_Izquierdo())
        self.ui.AjustesBtn.clicked.connect(self.Ajustes)
        self.ui.AyudaBtn.clicked.connect(self.Ayuda)
        self.ui.UserBtn.clicked.connect(lambda: self.Usuarios())
        self.ui.InterRedBtn.clicked.connect(lambda: self.Dipositivos())
        self.ui.AlertasBtn.clicked.connect(lambda: self.Alerta())
        
        self.ui.IPConfBtn.clicked.connect(lambda:self.IPCONFIG())
        self.ui.PingBtn.clicked.connect(lambda:self.PING())
        self.ui.ArpBtn.clicked.connect(lambda: self.ARP())
        self.ui.TracerBtn.clicked.connect(lambda:self.TRACERT())
        self.ui.RouteBtn.clicked.connect(lambda:self.ROUTE())
        self.ui.IPConfBtn.clicked.connect(lambda:self.IPCONFIG())
        self.ui.NSLookBtn.clicked.connect(lambda:self.NS())

        self.ui.AplicarAjusBtn.clicked.connect(self.AplicarConf)


        ##############################################
        ## BOTONES PAGINAS
        ##############################################
        self.ui.VelocidadBtn.clicked.connect(self.Velocidad)
        self.ui.DashboardBtn.clicked.connect(self.DashBoard)
        self.ui.GestionBtn.clicked.connect(self.Gestion)
        self.ui.TraficoBtn.clicked.connect(self.Trafico)
        self.ui.UtilidadesBtn.clicked.connect(self.Utilidades)
        self.ui.RerportesBtn.clicked.connect(self.Reportes)
        self.ui.BasuraNotif.clicked.connect(self.BorrarNo)
        ######################## Menus
        self.ui.DerechoBtn.clicked.connect(self.cerrar_MenuD)
        self.ui.MasMenuBtn.clicked.connect(self.cerrar_MenuC)
        self.ui.CerrarNotif.clicked.connect(self.cerrar_MenuN)

        self.ui.FiltradoSpeed.textChanged.connect(self.FiltrarSpeed)
        self.ui.Filtrado.textChanged.connect(self.FiltrarPaquete)
        ##Funcionalidades
        self.ui.PruebaVelBtn.clicked.connect(self.pruebaVel)  
        self.ui.PruebaVelTBtn.clicked.connect(self.pruebaVelT)  
        self.ui.BorrarPruebas.clicked.connect(self.borrarVel)
        #Trafico de paquetes
        self.ui.InicioTraficoBtn.clicked.connect(self.trafico)
        ############## Botones Gestion

        self.ui.G_PR.clicked.connect(self.DBproblemasRed)
        self.ui.G_R.clicked.connect(self.DBreportes)
        self.ui.G_I.clicked.connect(self.DBinvernario)
        self.ui.G_P.clicked.connect(self.DBProveedores)
        self.ui.G_RD.clicked.connect(self.DBRedes)
        self.ui.G_C.clicked.connect(self.DBConexiones)
        self.ui.G_H.clicked.connect(self.DBHosts)
        self.ui.G_CR.clicked.connect(self.DBControlRed)
        self.ui.G_SU.clicked.connect(self.DBUnidades)
        self.ui.G_D.clicked.connect(self.DBDepartamentos)
        self.ui.GAgregarBtn.clicked.connect(self.DBagregar)
        self.ui.GModificarBtn.clicked.connect(self.DBModificar)
        self.ui.GEliminarBtn.clicked.connect(self.DBEliminar)

        self.ui.CR_ADD.clicked.connect(self.DBCR_ADD)
        self.ui.D_ADD.clicked.connect(self.DBD_ADD)
        self.ui.I_ADD.clicked.connect(self.DBI_ADD)
        self.ui.P_ADD.clicked.connect(self.DBP_ADD)
        self.ui.G_AgregarBtn.clicked.connect(self.DBP_PR)
        self.ui.R_ADD.clicked.connect(self.DBP_R)
        self.ui.RD_ADD.clicked.connect(self.DBP_RD)
        self.ui.SU_ADD.clicked.connect(self.DBP_SU)
        
        self.ui.GFiltrado.textChanged.connect(self.Filtrar)
        ############### Utilidades
        self.ui.PingIniciarBtn.clicked.connect(self.utilidadPingIniciar)
        self.ui.PingAyudaBtn.clicked.connect(self.utilidadPingAyuda)

        self.ui.Iconfig1.clicked.connect(self.IPCAll)
        self.ui.Iconfig2.clicked.connect(self.IPCDDNS)
        self.ui.Iconfig3.clicked.connect(self.IPCFDNS)
        self.ui.Iconfig4.clicked.connect(self.IPCRDNS)
        self.ui.Iconfig5.clicked.connect(self.IPCAYUDA)
        self.ui.Iconfig6.clicked.connect(self.IPCMANUAL)
        
        self.ui.arp1.clicked.connect(self.ARPA)
        self.ui.arp2.clicked.connect(self.ARPC)
        self.ui.arp5.clicked.connect(self.ARPADD)
        self.ui.arp7.clicked.connect(self.ARPR)
        self.ui.ArpAyuda.clicked.connect(self.ARPAY)

        self.ui.NSLOIniciarBtn.clicked.connect(self.NSLOI)
        self.ui.NSLOAyudaBtn.clicked.connect(self.NSLOA)

        self.ui.TracertAyudaBtn.clicked.connect(self.tracertA)
        self.ui.TracertIniciarBtn.clicked.connect(self.tracertI)

        self.ui.var1.clicked.connect(self.VMac)
        self.ui.var2.clicked.connect(self.VARP)
        self.ui.var3.clicked.connect(self.speedtest)
        self.ui.var4.clicked.connect(self.fast)
        self.ui.var5.clicked.connect(self.MRed)
        self.ui.var6.clicked.connect(self.MTareas)
        self.ui.var7.clicked.connect(self.RCompartidos)
        self.ui.var8.clicked.connect(self.cmd)
        ############### Reportes
        self.ui.GRG.clicked.connect(self.RGeneral)
        self.ui.GR1.clicked.connect(self.RVelocidad)
        self.ui.GR2.clicked.connect(self.RServicio)
        self.ui.GR3.clicked.connect(self.RClues)
        self.ui.GR4.clicked.connect(self.RDisp)
        self.ui.GR5.clicked.connect(self.RHistorial)
        self.ui.GR6.clicked.connect(self.RProveedores)
        self.ui.GR7.clicked.connect(self.RInventario)
        self.ui.GR8.clicked.connect(self.RProblemas)

        self.ui.GRBorrarH.clicked.connect(self.BRHistorial)
        self.ui.GRBorrarS.clicked.connect(self.BRServicio)
        self.ui.GRBorrarP.clicked.connect(self.BRPruebas)
        self.ui.GRBorrarPr.clicked.connect(self.BRProblemas)


        ############### Trafico        
        self.ui.VerDomBtn.clicked.connect(self.dominios)
        self.ui.SalirSalidaPaq.clicked.connect(self.salirPaq)
        self.ui.RestabBtn.clicked.connect(self.BorrarTabla)
        ###############Señales
        self.signalUser.connect(self.UsuariosDisp)
        self.signalPing.connect(self.putPing)
        self.signalTrafico.connect(self.FiltrarPaquete)
        self.SignalV.connect(self.putVelocidad)
        self.SignalVT.connect(self.putVelocidadT)
        self.SignalVTL.connect(self.pruebaCompletaVT)
        self.senalS.connect(self.crearGraficaSignal)
        self.SignalA.connect(self.SignalAlerta)
        self.tab.connect(self.crearTablas)
        ##############Modificar tabla
        
        self.ui.PINGOUT.setColumnWidth(0,50)
        self.ui.PINGOUT.setColumnWidth(1,100)
        self.ui.PINGOUT.setColumnWidth(2,100)
        self.ui.PINGOUT.setColumnWidth(3,70)
        self.ui.PINGOUT.setColumnWidth(4,70)

        self.ui.TablaPaquetes.doubleClicked.connect(self.Doubleclick)
        self.ui.SalidaGestion.doubleClicked.connect(self.dobleGestion)
        self.ui.TablaVelocidad.doubleClicked.connect(self.dobleVelocidad)
        
        self.ui.Ajus4.textChanged.connect(lambda text: self.validarNumero(self.ui.Ajus4))
        self.ui.Ajus5.textChanged.connect(lambda text: self.validarNumero(self.ui.Ajus5))
        self.ui.RD_DAJ.textChanged.connect(lambda text: self.validarNumero(self.ui.RD_DAJ))
        self.ui.RD_SUB.textChanged.connect(lambda text: self.validarNumero(self.ui.RD_SUB))
        self.ui.ping3.textChanged.connect(lambda text: self.validarNumero(self.ui.ping3))
        self.ui.ping12.textChanged.connect(lambda text: self.validarNumero(self.ui.ping12))
        self.ui.ping2.textChanged.connect(lambda text: self.validarNumero(self.ui.ping2))
        self.ui.ping4.textChanged.connect(lambda text: self.validarNumero(self.ui.ping4))
        self.ui.tracert1.textChanged.connect(lambda text: self.validarNumero(self.ui.tracert1))
        self.ui.tracert2.textChanged.connect(lambda text: self.validarNumero(self.ui.tracert2))
        self.actualizarDatosDash()
        self.crearTablas()
        
        self.DBproblemasRed()
        disp = QVBoxLayout()
        disp.addWidget(QLabel(f"<p align='center'><b>Buscando Dispositivos en {self.inf.SSID}...</b></p>")) 
        aux2=QWidget()
        aux2.setLayout(disp)
        aux2.setContentsMargins(0,0,0,0)
        self.ui.scrollArea_2.setWidget(aux2)
        self.ui.label_5.setText(f"<p align='right'><b>Red:</b> {self.inf.SSID}&nbsp; &nbsp; &nbsp; <b>IPv4:</b> {self.inf.IPV4}&nbsp; &nbsp; &nbsp; <b>Gateway:</b> {self.inf.GATEWAY}</p>")

        self.mostrarErrores()
       
        Usuarios = Thread(target=self.hilo_usuarios)
        Usuarios.start()
        RedCambios = Thread(target=self.CambiosRed)
        RedCambios.start()

        sen=Thread(target=self.senal)
        sen.start()
        self.ui.SpeedRed.setText(self.inf.SSID)

    velocidad=None
    def dobleVelocidad(self):
        row_number=0
        for idx in self.ui.TablaVelocidad.selectionModel().selectedIndexes():
            row_number = idx.row()
        aux=[]
        for i in range(0,self.ui.TablaVelocidad.columnCount()):
            aux.append(self.ui.TablaVelocidad.item(row_number,i).text())
        r=db.consultar(f"SELECT ID_RED,SUBIDA,BAJADA FROM RED WHERE SSID='{aux[5]}'")
        subida=r[0][1]  
        bajada=r[0][2]
        SubidaPor=float(aux[6])*100/subida
        BajadaPor=float(aux[7])*100/bajada
        if self.inf.TEMA=="Obscuro":
            BIEN="QFrame{background-color:#282a36; border:10px solid; border-color:#02AC66;color:#f8f8f2;border-radius:100px;} QLabel{Border:0px}"
            MEDIO="QFrame{background-color:#282a36; border:10px solid; border-color:#024A86;color:#f8f8f2;border-radius:100px;} QLabel{Border:0px}"
            MAL="QFrame{background-color:#282a36; border:10px solid; border-color:#C82A54;color:#f8f8f2;border-radius:100px;} QLabel{Border:0px}"
        else:
            BIEN="QFrame{background-color:#D8D8D8; border:10px solid; border-color:#02AC66;color:#333333;border-radius:100px;} QLabel{Border:0px;color:#333333;}"
            MEDIO="QFrame{background-color:#D8D8D8; border:10px solid; border-color:#024A86;color:#333333;border-radius:100px;} QLabel{Border:0px;color:#333333;}"
            MAL="QFrame{background-color:#D8D8D8; border:10px solid; border-color:#C82A54;color:#333333;border-radius:100px;} QLabel{Border:0px;color:#333333;}"
        self.ui.Subida_Salida.setText(aux[6]+" Mbps")
        if SubidaPor>90: self.ui.Subida_C.setStyleSheet(BIEN)
        elif SubidaPor<90 and SubidaPor>50: self.ui.Subida_C.setStyleSheet(MEDIO)
        elif SubidaPor>50: self.ui.Subida_C.setStyleSheet(MAL)
        self.ui.Velocidad_Salida.setText(aux[7]+" Mbps")
        if BajadaPor>90: self.ui.Velocidad_C.setStyleSheet(BIEN)
        elif BajadaPor<90 and BajadaPor>50: self.ui.Velocidad_C.setStyleSheet(MEDIO)
        elif BajadaPor>50: self.ui.Velocidad_C.setStyleSheet(MAL)
        self.ui.Ping_Salida.setText(aux[10]+" ms")
        if float(aux[10])<50: self.ui.Ping_C.setStyleSheet(BIEN)
        elif float(aux[10])<100 and float(aux[10])>50: self.ui.Ping_C.setStyleSheet(MEDIO)
        elif float(aux[10])>100: self.ui.Ping_C.setStyleSheet(MAL)

        self.ui.SpeedRed.setText(aux[5])
        self.ui.SpeedBajada.setText(aux[9]+" ms")
        self.ui.SpeedSubida.setText(aux[8]+" ms")
        self.ui.SpeedServidor.setText(aux[3])
        self.ui.SpeedSponsor.setText(aux[4])
    def borrarVel(self):
        db.ejecutarAccion("DELETE FROM PRUEBAS_VELOCIDAD")
        self.DBPruebas()            
        self.inf.Errores.append([f"{dt.now().hour}:{dt.now().minute}","Pruebas de Velocidad","Se han borrado las pruebas de red."])
    def pruebaVel(self):
        self.ui.PruebaVelBtn.setStyleSheet("border-radius:10px;background-color: #20945e;	padding: 8px 8px; color:#FFFFFF;")
        self.ui.PruebaVelBtn.setText("Espera un momento")
        self.ui.PruebaVelBtn.setEnabled(False)
        self.ui.PruebaVelTBtn.setEnabled(False)
        
        hilo=Thread(target=self.hiloVel)
        hilo.start()
    def pruebaVelT(self):
        self.ui.PruebaVelTBtn.setStyleSheet("border-radius:10px;background-color: #20945e;	padding: 8px 8px; color:#FFFFFF;")
        self.ui.PruebaVelTBtn.setText("Espera un momento")
        self.ui.PruebaVelTBtn.setEnabled(False)
        self.ui.PruebaVelBtn.setEnabled(False)
        hilo=Thread(target=self.hiloVelT)
        hilo.start()
    def hiloVelT(self):
        try:
            self.velocidad=sd.velocidad_Todo(self.TrafObj,self.SignalVT,self.inf,self.SignalVTL)
        except:
            self.inf.Errores.append([f"{dt.now().hour}:{dt.now().minute}","Conexión red.","Hay un error general"])
        self.mostrarErrores()
        self.ui.PruebaVelTBtn.setEnabled(True)
        self.ui.PruebaVelBtn.setEnabled(True)
    def hiloVel(self):
        try:
            self.velocidad=sd.Velocidad_Internet(self.inf)
            self.SignalV.emit(1)
        except:
            self.inf.Errores.append([f"{dt.now().hour}:{dt.now().minute}",self.inf.SSID,"No hay internet en la red para la prueba."])
            if self.inf.TEMA=="Obscuro":
                self.ui.PruebaVelBtn.setStyleSheet("border-radius:10px;background-color: #1f232a;	padding: 8px 8px;")
            else:
                self.ui.PruebaVelBtn.setStyleSheet("border-radius:10px;background-color:#D8D8D8; padding: 8px 8px;")
            self.ui.PruebaVelBtn.setText("Iniciar")
            self.ui.PruebaVelTBtn.setEnabled(True)
            self.ui.PruebaVelBtn.setEnabled(True)
        self.mostrarErrores()

    def putVelocidadT(self):
        
        now = dt.now()
        r=db.consultar(f"SELECT ID_RED,SUBIDA,BAJADA FROM RED WHERE SSID='{self.inf.SSID}'")
        red=r[0][0]
        subida=r[0][1]  
        bajada=r[0][2]
        self.ui.Velocidad_Salida.setText(str(self.TrafObj.velocidad[0])+" Mbps")
        self.ui.Subida_Salida.setText(str(self.TrafObj.velocidad[1])+" Mbps")
        self.ui.Ping_Salida.setText(str(self.TrafObj.velocidad[4])+ " ms")
        if subida==0:subida=float(self.TrafObj.velocidad[1])
        if bajada==0:  bajada=float(self.TrafObj.velocidad[0])
        SubidaPor=float(self.TrafObj.velocidad[1])*100/subida
        BajadaPor=float(self.TrafObj.velocidad[0])*100/bajada
        Porcentaje=(SubidaPor+BajadaPor)/2
        Ping=float(self.TrafObj.velocidad[4])
        if self.inf.TEMA=="Obscuro":
            BIEN="QFrame{background-color:#282a36; border:10px solid; border-color:#02AC66;color:#f8f8f2;border-radius:100px;} QLabel{Border:0px}"
            MEDIO="QFrame{background-color:#282a36; border:10px solid; border-color:#024A86;color:#f8f8f2;border-radius:100px;} QLabel{Border:0px}"
            MAL="QFrame{background-color:#282a36; border:10px solid; border-color:#C82A54;color:#f8f8f2;border-radius:100px;} QLabel{Border:0px}"
        else:
            BIEN="QFrame{background-color:#D8D8D8; border:10px solid; border-color:#02AC66;color:#333333;border-radius:100px;} QLabel{Border:0px;color:#333333;}"
            MEDIO="QFrame{background-color:#D8D8D8; border:10px solid; border-color:#024A86;color:#333333;border-radius:100px;} QLabel{Border:0px;color:#333333;}"
            MAL="QFrame{background-color:#D8D8D8; border:10px solid; border-color:#C82A54;color:#333333;border-radius:100px;} QLabel{Border:0px;color:#333333;}"
        if SubidaPor>=80: self.ui.Subida_C.setStyleSheet(BIEN)
        elif SubidaPor<80 and SubidaPor>=50: self.ui.Subida_C.setStyleSheet(MEDIO)
        elif SubidaPor<50: self.ui.Subida_C.setStyleSheet(MAL)
        
        if BajadaPor>=80: self.ui.Velocidad_C.setStyleSheet(BIEN)
        elif BajadaPor<80 and BajadaPor>=50: self.ui.Velocidad_C.setStyleSheet(MEDIO)
        elif BajadaPor<50: self.ui.Velocidad_C.setStyleSheet(MAL)
        
        if Ping<=50: self.ui.Ping_C.setStyleSheet(BIEN)
        elif Ping<=100 and BajadaPor>50: self.ui.Ping_C.setStyleSheet(MEDIO)
        elif Ping>100: self.ui.Ping_C.setStyleSheet(MAL)

        Estado=""
        if Porcentaje>=95:Estado="Excelente"
        elif Porcentaje>=70 and Porcentaje<95:Estado="Bueno"
        elif Porcentaje>=40 and Porcentaje<70:Estado="Regular"
        elif Porcentaje<40:Estado="Malo"
        self.ui.SpeedBajada.setText(self.TrafObj.velocidad[2]+" ms")
        self.ui.SpeedSubida.setText(self.TrafObj.velocidad[3]+" ms")
        self.ui.SpeedServidor.setText(self.TrafObj.velocidad[5])
        self.ui.SpeedSponsor.setText(self.TrafObj.velocidad[6].replace("ISP:",""))
        db.insertarPRUEBAS_VELOCIDAD(red,f"{now.year}-{now.month}-{now.day}",f"{now.hour}:{now.minute}",self.TrafObj.velocidad[5],self.TrafObj.velocidad[6].replace("ISP:",""),float(self.TrafObj.velocidad[1]),float(self.TrafObj.velocidad[0]),float(self.TrafObj.velocidad[2]),self.TrafObj.velocidad[3],float(self.TrafObj.velocidad[4]),Estado)
        self.DBPruebas()            
    def pruebaCompletaVT(self):
        if self.inf.TEMA=="Obscuro":
            self.ui.PruebaVelTBtn.setStyleSheet("border-radius:10px;background-color: #1f232a;	padding: 8px 8px;")
        else:
            self.ui.PruebaVelTBtn.setStyleSheet("border-radius:10px;background-color: #FFFFFF;	padding: 8px 8px;")
        self.ui.PruebaVelTBtn.setText("Iniciar para todas las redes.(Wi-Fi)")
        self.ui.PruebaVelTBtn.setEnabled(True)
        self.ui.PruebaVelTBtn.setEnabled(True)
    def putVelocidad(self):
        now = dt.now()
        r=db.consultar(f"SELECT ID_RED,SUBIDA,BAJADA FROM RED WHERE SSID='{self.inf.SSID}'")
        red=r[0][0]
        subida=r[0][1]  
        bajada=r[0][2]
        self.ui.Velocidad_Salida.setText(str(self.velocidad[0])+" Mbps")
        self.ui.Subida_Salida.setText(str(self.velocidad[1])+" Mbps")
        self.ui.Ping_Salida.setText(str(self.velocidad[4])+ " ms")
        if subida==0:subida=float(self.velocidad[1])
        if bajada==0:  bajada=float(self.velocidad[0])
        SubidaPor=float(self.velocidad[1])*100/subida
        BajadaPor=float(self.velocidad[0])*100/bajada
        Porcentaje=(SubidaPor+BajadaPor)/2
        Ping=float(self.velocidad[4])
        if self.inf.TEMA=="Obscuro":
            BIEN="QFrame{background-color:#282a36; border:10px solid; border-color:#02AC66;color:#f8f8f2;border-radius:100px;} QLabel{Border:0px}"
            MEDIO="QFrame{background-color:#282a36; border:10px solid; border-color:#024A86;color:#f8f8f2;border-radius:100px;} QLabel{Border:0px}"
            MAL="QFrame{background-color:#282a36; border:10px solid; border-color:#C82A54;color:#f8f8f2;border-radius:100px;} QLabel{Border:0px}"
        else:
            BIEN="QFrame{background-color:#D8D8D8; border:10px solid; border-color:#02AC66;color:#f8f8f2;border-radius:100px;} QLabel{Border:0px;color:#333333;}"
            MEDIO="QFrame{background-color:#D8D8D8; border:10px solid; border-color:#024A86;color:#f8f8f2;border-radius:100px;} QLabel{Border:0px;color:#333333;}"
            MAL="QFrame{background-color:#D8D8D8; border:10px solid; border-color:#C82A54;color:#f8f8f2;border-radius:100px;} QLabel{Border:0px;color:#333333;}"
        if SubidaPor>=80: self.ui.Subida_C.setStyleSheet(BIEN)
        elif SubidaPor<80 and SubidaPor>=50: self.ui.Subida_C.setStyleSheet(MEDIO)
        elif SubidaPor<50: self.ui.Subida_C.setStyleSheet(MAL)
        
        if BajadaPor>=80: self.ui.Velocidad_C.setStyleSheet(BIEN)
        elif BajadaPor<80 and BajadaPor>=50: self.ui.Velocidad_C.setStyleSheet(MEDIO)
        elif BajadaPor<50: self.ui.Velocidad_C.setStyleSheet(MAL)
        
        if Ping<=50: self.ui.Ping_C.setStyleSheet(BIEN)
        elif Ping<=100 and BajadaPor>50: self.ui.Ping_C.setStyleSheet(MEDIO)
        elif Ping>100: self.ui.Ping_C.setStyleSheet(MAL)

        Estado=""
        if Porcentaje>=95:Estado="Excelente"
        elif Porcentaje>=70 and Porcentaje<95:Estado="Bueno"
        elif Porcentaje>=40 and Porcentaje<70:Estado="Regular"
        elif Porcentaje<40:Estado="Malo"
        self.ui.SpeedBajada.setText(self.velocidad[2]+" ms")
        self.ui.SpeedSubida.setText(self.velocidad[3]+" ms")
        self.ui.SpeedServidor.setText(self.velocidad[5])
        self.ui.SpeedSponsor.setText(self.velocidad[6].replace("ISP:",""))
        db.insertarPRUEBAS_VELOCIDAD(red,f"{now.year}-{now.month}-{now.day}",f"{now.hour}:{now.minute}",self.velocidad[5],self.velocidad[6].replace("ISP:",""),float(self.velocidad[1]),float(self.velocidad[0]),float(self.velocidad[2]),self.velocidad[3],float(self.velocidad[4]),Estado)
        self.DBPruebas()
        if self.inf.TEMA=="Obscuro":            
            self.ui.PruebaVelBtn.setStyleSheet("border-radius:10px;background-color: #1f232a;	padding: 8px 8px;")
        else:
            self.ui.PruebaVelBtn.setStyleSheet("border-radius:10px;background-color: #FFFFFF;	padding: 8px 8px;")
        self.ui.PruebaVelBtn.setText("Iniciar")
        self.ui.PruebaVelBtn.setEnabled(True)
        self.ui.PruebaVelTBtn.setEnabled(True)
    ##############################################
    ## Encabezado
    ##############################################
    def AplicarConf(self):
        aux1=self.ui.Ajus1.itemText(self.ui.Ajus1.currentIndex())
        aux2=self.ui.Ajus2.itemText(self.ui.Ajus2.currentIndex())
        aux3=self.ui.Ajus3.itemText(self.ui.Ajus3.currentIndex())
        aux4=self.ui.Ajus4.text()
        aux5=self.ui.Ajus5.text()
        file=open("modulos/BasesDatos/configuracion.conf","w",encoding="utf-8")
        file.write(f"{aux1}\n")
        file.write(f"{aux2}\n")
        file.write(f"{aux3}\n")
        file.write(f"{aux4}\n")
        file.write(f"{aux5}")
        file.close()
        self.inf.SALIR=True
        self.TrafObj.Cerrar=True
        self.close()
        os.execl(sys.executable, "main.py", *sys.argv) 


    def mouseDoubleClickEvent(self, event):
        widget = self.childAt(event.pos())
        if widget is not None and widget.objectName() == "Encabezado":
            self.Maximizar_Restaurar()
    def validarNumero(self,text):
        aux=text.text()
        try:
            float(aux)
        except:
            text.setText(aux[0:len(aux)-1])
    def Cerrar(self):
        self.inf.SALIR=True
        self.TrafObj.Cerrar=True
        self.close()
        sys.exit(sys.executable)
    def closeEvent(self, event):
        self.inf.SALIR=True
        self.TrafObj.Cerrar=True
        event.accept()
    def Minimizar(self):
        self.showMinimized()
    def Maximizar_Restaurar(self):
        status=self.GLOBAL_STATE
        if status == False:
            self.showMaximized()
            self.GLOBAL_STATE = True
            self.ui.MaxBtn.setToolTip("Restaurar")
            self.ui.MaxBtn.setIcon(QIcon(u":/Iconos2/Imagenes/icons/icon_maximize.png"))
        else:
            self.GLOBAL_STATE = False
            self.showNormal()
            self.resize(1089, 640)
            self.ui.MaxBtn.setToolTip("Maximizar")
            self.ui.MaxBtn.setIcon(QIcon(u":/Iconos2/Imagenes/icons/cil-window-maximize.png"))
        self.ui.SalidaGestion.resizeRowsToContents()
        for i in range(0,self.ui.SalidaGestion.columnCount()): self.ui.SalidaGestion.setColumnWidth(i,self.ui.SalidaGestion.width()/self.ui.SalidaGestion.columnCount())
        self.ui.TablaVelocidad.resizeRowsToContents()
        for i in range(0,self.ui.TablaVelocidad.columnCount()): self.ui.TablaVelocidad.setColumnWidth(i,self.ui.TablaVelocidad.width()/self.ui.TablaVelocidad.columnCount())
        self.ui.TablaPaquetes.resizeRowsToContents()
        for i in range(0,self.ui.TablaPaquetes.columnCount()): self.ui.TablaPaquetes.setColumnWidth(i,self.ui.TablaPaquetes.width()/self.ui.TablaPaquetes.columnCount())
        self.ui.PINGOUT.resizeRowsToContents()
        for i in range(0,self.ui.PINGOUT.columnCount()): self.ui.PINGOUT.setColumnWidth(i,self.ui.PINGOUT.width()/self.ui.PINGOUT.columnCount())
    pressing = False
    def mousePressEvent(self, event):
        widget = self.childAt(event.pos())
        if widget is not None and widget.objectName() == "Encabezado":
            self.dragPosition = event.globalPos() - self.frameGeometry().topLeft()
            event.accept()
            self.pressing = True
    def mouseMoveEvent(self, event):
        if self.pressing:
            if self.GLOBAL_STATE==True:
                self.Maximizar_Restaurar()
                self.GLOBAL_STATE==False
                self.move( self.dragPosition)
            self.move(event.globalPos() - self.dragPosition)
            event.accept()
    def mouseReleaseEvent(self, QMouseEvent):
        self.pressing = False
    Side_menu_I=0
    def animacion_lateral_Inicio(self,p1,p2,t,elemento):
            self.animacion = QPropertyAnimation(elemento,b"maximumWidth")
            self.animacion.setDuration(t)
            self.animacion.setStartValue(p1)
            self.animacion.setEndValue(p2)
            self.animacion.setEasingCurve(QEasingCurve.InOutQuart)
            self.animacion.start()
        
            self.animacion2 = QPropertyAnimation(elemento,b"minimumWidth")
            self.animacion2.setDuration(t)
            self.animacion2.setStartValue(p1)
            self.animacion2.setEndValue(p2)
            self.animacion2.setEasingCurve(QEasingCurve.InOutQuart)
            self.animacion2.start()
            
    def animacion_lateral_Fin(self,p1,p2,t,elemento):
            self.animacion = QPropertyAnimation(elemento,b"minimumWidth")
            self.animacion.setDuration(t)
            self.animacion.setStartValue(p2)
            self.animacion.setEndValue(p1)
            self.animacion.setEasingCurve(QEasingCurve.InOutQuart)
            self.animacion.start()
        
            self.animacion2 = QPropertyAnimation(elemento,b"maximumWidth")
            self.animacion2.setDuration(t)
            self.animacion2.setStartValue(p2)
            self.animacion2.setEndValue(p1)
            self.animacion2.setEasingCurve(QEasingCurve.InOutQuart)
            self.animacion2.start()
    def animacion_Vertical_Inicio(self,p1,p2,t,elemento):
            self.animacion = QPropertyAnimation(elemento,b"maximumHeight")
            self.animacion.setDuration(t)
            self.animacion.setStartValue(p1)
            self.animacion.setEndValue(p2)
            self.animacion.setEasingCurve(QEasingCurve.InOutQuart)
            self.animacion.start()
        
            self.animacion2 = QPropertyAnimation(elemento,b"minimumHeight")
            self.animacion2.setDuration(t)
            self.animacion2.setStartValue(p1)
            self.animacion2.setEndValue(p2)
            self.animacion2.setEasingCurve(QEasingCurve.InOutQuart)
            self.animacion2.start()
            
    def animacion_Vertical_Fin(self,p1,p2,t,elemento):
            self.animacion = QPropertyAnimation(elemento,b"minimumHeight")
            self.animacion.setDuration(t)
            self.animacion.setStartValue(p2)
            self.animacion.setEndValue(p1)
            self.animacion.setEasingCurve(QEasingCurve.InOutQuart)
            self.animacion.start()
        
            self.animacion2 = QPropertyAnimation(elemento,b"maximumHeight")
            self.animacion2.setDuration(t)
            self.animacion2.setStartValue(p2)
            self.animacion2.setEndValue(p1)
            self.animacion2.setEasingCurve(QEasingCurve.InOutQuart)
            self.animacion2.start()
    def Menu_Izquierdo(self):
        if self.Side_menu_I==0:
            self.animacion_lateral_Inicio(45,145,500,self.ui.MenuIzquierdo)
            self.Side_menu_I=1
        else:
            self.animacion_lateral_Fin(45,145,500,self.ui.MenuIzquierdo)
            self.Side_menu_I=0
    def cerrar_MenuC(self):
        if self.inf.TEMA=="Obscuro":
            style="QPushButton{text-align:left;	padding:5px 10px;	border-top-left-radius: 10px;	border-bottom-left-radius: 10px;} QPushButton::hover{background-color:#1f232a;}QPushButton::pressed{background-color:#343b47;}"
        else:
            style="QPushButton{text-align:left;	padding:5px 10px;	border-top-left-radius: 10px;	border-bottom-left-radius: 10px;} QPushButton::hover{background-color:#FFFFFF;}QPushButton::pressed{background-color:#FFFFFF;}"
        self.ui.AjustesBtn.setStyleSheet(style)
        self.ui.AyudaBtn.setStyleSheet(style)
        self.animacion_lateral_Fin(0,230,500,self.ui.MenuCentral)
        self.Side_menu_A=0
    Side_menu_H=0 
    def Ayuda(self):
        if self.Side_menu_H==0 and self.Side_menu_A == 0:
            style="QPushButton{text-align:left;	padding:5px 10px;	border-top-left-radius: 10px;	border-bottom-left-radius: 10px;} QPushButton::hover{background-color:#1f232a;}QPushButton::pressed{background-color:#343b47;}"
            self.ui.AjustesBtn.setStyleSheet(style)
            style="QPushButton{background-color:#1f232a; text-align:left;	padding:5px 10px;	border-top-left-radius: 10px;	border-bottom-left-radius: 10px;} QPushButton::hover{background-color:#1f232a;}QPushButton::pressed{background-color:#343b47;}"
            self.ui.AyudaBtn.setStyleSheet(style)
            self.ui.ContenidoCentral.setCurrentWidget(self.ui.pAyu) 
            self.animacion_lateral_Inicio(0,300,500,self.ui.MenuCentral)
            self.ui.label.setText("<b>Ayuda</b>")
            self.Side_menu_H=1
            os.system("START modulos/BasesDatos/Manual.pdf")

        elif self.Side_menu_A == 1 and self.Side_menu_H == 0:
            style="QPushButton{text-align:left;	padding:5px 10px;	border-top-left-radius: 10px;	border-bottom-left-radius: 10px;} QPushButton::hover{background-color:#1f232a;}QPushButton::pressed{background-color:#343b47;}"
            self.ui.AjustesBtn.setStyleSheet(style)
            style="QPushButton{background-color:#1f232a; text-align:left;	padding:5px 10px;	border-top-left-radius: 10px;	border-bottom-left-radius: 10px;} QPushButton::hover{background-color:#1f232a;}QPushButton::pressed{background-color:#343b47;}"
            self.ui.AyudaBtn.setStyleSheet(style)
            self.ui.ContenidoCentral.setCurrentWidget(self.ui.pAyu) 
            self.Side_menu_A=0
            self.Side_menu_H=1
            self.ui.label.setText("<b>Ayuda</b>")
            os.system("START modulos/BasesDatos/Manual.pdf")

        else:
            style="QPushButton{text-align:left;	padding:5px 10px;	border-top-left-radius: 10px;	border-bottom-left-radius: 10px;} QPushButton::hover{background-color:#1f232a;}QPushButton::pressed{background-color:#343b47;}"
            self.ui.AyudaBtn.setStyleSheet(style)
            self.animacion_lateral_Fin(0,300,500,self.ui.MenuCentral)
            self.Side_menu_H=0
    Side_menu_A=0 
    def menuAjus(self):
        aux=open("modulos/BasesDatos/configuracion.conf","r",encoding="utf-8")
        contenido=aux.readlines()
        self.inf.VENTANA=contenido[0].replace("\n","")
        self.inf.TEMA=contenido[1].replace("\n","")
        self.inf.DETECCION=contenido[2].replace("\n","")
        self.inf.TIEMPO_DETECCION=float(contenido[3].replace("\n",""))
        self.inf.TIEMPO_PAQ=float(contenido[4].replace("\n",""))
        aux.close()
        self.ui.Ajus1.clear()
        if self.inf.VENTANA=="Propia del Software":
            self.ui.Ajus1.addItem("Propia del Software")
            self.ui.Ajus1.addItem("Propia del SO")
        else:
            self.ui.Ajus1.addItem("Propia del SO")
            self.ui.Ajus1.addItem("Propia del Software")
        self.ui.Ajus2.clear()
        if self.inf.TEMA=="Obscuro":
            self.ui.Ajus2.addItem("Obscuro")
            self.ui.Ajus2.addItem("Claro")
        else:
            self.ui.Ajus2.addItem("Claro")
            self.ui.Ajus2.addItem("Obscuro")
        self.ui.Ajus3.clear()
        if self.inf.DETECCION=="Preciso (ARP y Ping)":
            self.ui.Ajus3.addItem("Preciso (ARP y Ping)")
            self.ui.Ajus3.addItem("Rápido (Solo ARP)")
        else:
            self.ui.Ajus3.addItem("Rápido (Solo ARP)")
            self.ui.Ajus3.addItem("Preciso (ARP y Ping)")
        self.ui.Ajus4.setText(str(self.inf.TIEMPO_DETECCION))
        self.ui.Ajus5.setText(str(self.inf.TIEMPO_PAQ))
    def Ajustes(self):
        if self.Side_menu_A==0 and self.Side_menu_H == 0:
            style="QPushButton{text-align:left;	padding:5px 10px;	border-top-left-radius: 10px;	border-bottom-left-radius: 10px;} QPushButton::hover{background-color:#1f232a;}QPushButton::pressed{background-color:#343b47;}"
            self.ui.AyudaBtn.setStyleSheet(style)

            style="QPushButton{background-color:#1f232a; text-align:left;	padding:5px 10px;	border-top-left-radius: 10px;	border-bottom-left-radius: 10px;} QPushButton::hover{background-color:#1f232a;}QPushButton::pressed{background-color:#343b47;}"
            self.ui.AjustesBtn.setStyleSheet(style)
            self.ui.ContenidoCentral.setCurrentWidget(self.ui.pAjus) 
            self.animacion_lateral_Inicio(0,300,500,self.ui.MenuCentral)
            
            self.Side_menu_A=1
            self.ui.BarDis.clear()
            self.menuAjus()
            self.ui.label.setText("<b>Ajustes</b>")
        elif self.Side_menu_H == 1 and self.Side_menu_A == 0:
            style="QPushButton{text-align:left;	padding:5px 10px;	border-top-left-radius: 10px;	border-bottom-left-radius: 10px;} QPushButton::hover{background-color:#1f232a;}QPushButton::pressed{background-color:#343b47;}"
            self.ui.AyudaBtn.setStyleSheet(style)
            style="QPushButton{background-color:#1f232a; text-align:left;	padding:5px 10px;	border-top-left-radius: 10px;	border-bottom-left-radius: 10px;} QPushButton::hover{background-color:#1f232a;}QPushButton::pressed{background-color:#343b47;}"
            self.ui.AjustesBtn.setStyleSheet(style)
            self.ui.ContenidoCentral.setCurrentWidget(self.ui.pAjus)
            self.Side_menu_H=0
            self.Side_menu_A=1
            self.ui.label.setText("<b>Ajustes</b>")
            self.menuAjus()

        else:
            style="QPushButton{text-align:left;	padding:5px 10px;	border-top-left-radius: 10px;	border-bottom-left-radius: 10px;} QPushButton::hover{background-color:#1f232a;}QPushButton::pressed{background-color:#343b47;}"
            self.ui.AjustesBtn.setStyleSheet(style)
            self.ui.AyudaBtn.setStyleSheet(style)
            self.animacion_lateral_Fin(0,300,500,self.ui.MenuCentral)
            self.Side_menu_A=0
            
    Side_menu_D=0 
    def Usuarios(self):
        if self.Side_menu_D==0 and self.Side_menu_US == 0:
            self.ui.label_2.setText("Usuarios Dentro de la Red")
            self.ui.ContenidoDerecho.setCurrentWidget(self.ui.pUse) 
            self.animacion_lateral_Inicio(0,320,500,self.ui.MenuDerecho)

            self.Side_menu_D=1
            

        elif self.Side_menu_US == 1 and self.Side_menu_D == 0:
            self.ui.ContenidoDerecho.setCurrentWidget(self.ui.pUse)
            self.Side_menu_D=1
            self.Side_menu_US=0
           
        else:
            self.animacion_lateral_Fin(0,320,500,self.ui.MenuDerecho)
            self.Side_menu_D=0
    def hilo_usuarios(self):
        while self.inf.SALIR!=True:
            if self.inf.CONEXION==True:
                try:
                    if self.inf.CAMBIOS == False:
                        self.NUM,dispositivos,ips_arp=rs.escanearARP(self.inf.GATEWAY+"/24",self.inf.INTERFAZ)
                        aux=self.inf.GATEWAY.split(".")
                        GATEWAY_PING=aux[0]+"."+aux[1]+"."+aux[2]+".0/24"
                    if self.inf.SALIR==True: break
                    if self.inf.DETECCION=="Preciso (ARP y Ping)":
                        if self.inf.CAMBIOS == False:
                            ips_ping=rs.pingda(GATEWAY_PING,ips_arp,self.inf)
                        if self.inf.SALIR==True: break
                        if self.inf.CAMBIOS == False:
                            disposi=rs.ping_arp(ips_ping, ips_arp,dispositivos,self.inf.INTERFAZ)
                        if self.inf.SALIR==True: break
                    else:
                        disposi=dispositivos
                    if self.inf.CAMBIOS == False:
                        self.NUM,self.dispo=rs.obtener_datos_Dispositivo(disposi,self.inf.INTERFAZ,self.inf)
                    if self.inf.SALIR==True: break
                    if self.inf.CAMBIOS == False:
                        self.signalUser.emit(1)
                    self.inf.CAMBIOS=False
                    self.mostrarErrores()
                    sleep(self.inf.TIEMPO_DETECCION)
                except Exception as e:
                    print(f"[!] Error:{e}")
                    self.inf.Errores.append([f"{dt.now().hour}:{dt.now().minute}","Error de Aplicación","Se detecto el error"+str(type(e))])
                    sleep(10)
            else:
                sleep(1)
    def UsuariosDisp(self):
        disp = QVBoxLayout()
        disp.addWidget(QLabel("<p align='center'><b>Existen "+str(self.NUM)+" dispositivos conectados.</b></p>"))
        for NE in self.dispo:
            aux=UsuariosClass(NE,self.inf.TEMA,self.inf)
            disp.addWidget(aux) 
        aux2=QWidget()
        aux2.setLayout(disp)
        aux2.setContentsMargins(0,0,0,0)
        self.ui.scrollArea_2.setWidget(aux2)
        try:
            ID=int(db.consultar(f"SELECT ID_RED FROM RED WHERE SSID='{self.inf.SSID}'")[0][0])
            val=db.consultar(f"SELECT count(*) FROM CONTROL_RED_ESPECIFICA WHERE ID_RED={ID}")[0][0]
            self.ui.UserBtn.setText(f" {self.NUM} / {str(int(val))}")
        except:
            print("[!] Error al actualizar usuarios.")
            self.ui.UserBtn.setText("")
    Side_menu_US=0
    def Dipositivos(self):
        if self.Side_menu_US==0 and self.Side_menu_D == 0:
            #self.ui.SalidaARP.setHtml("<font color='cyan'>asasd</font>")
            self.ui.label_2.setText("Interfaces de red")
            self.ui.ContenidoDerecho.setCurrentWidget(self.ui.pDisp) 
            self.animacion_lateral_Inicio(0,320,500,self.ui.MenuDerecho)
            self.DispositivosFrame()
            self.Side_menu_US=1
        elif self.Side_menu_D == 1 and self.Side_menu_US == 0:
            self.ui.ContenidoDerecho.setCurrentWidget(self.ui.pDisp)
            self.DispositivosFrame()
            self.Side_menu_D=0
            self.Side_menu_US=1
        else:
            self.animacion_lateral_Fin(0,320,500,self.ui.MenuDerecho)
            self.Side_menu_US=0
    def DispositivosFrame(self):
        disp = QVBoxLayout()
        INTER=rs.Obtener_Dispositivos(True)
        for IN in INTER:
            aux=DispositivosClass(IN,self.inf.TEMA,self.inf)
            disp.addWidget(aux) 
        aux=QWidget()
        aux.setLayout(disp)
        aux.setContentsMargins(0,0,0,0)
        self.ui.scrollArea.setWidget(aux)
    def cerrar_MenuD(self):
        self.animacion_lateral_Fin(0,320,500,self.ui.MenuDerecho)
        self.Side_menu_US=0
    Side_menu_AL=0 
    def Alerta(self):
        self.formato(self.ui.AlertasBtn)
        if self.Side_menu_AL==0:
            self.animacion_Vertical_Inicio(0,200,500,self.ui.Notificacion)
            self.Side_menu_AL=1
        else:
            style="QPushButton{text-align:left;	padding:5px 10px;	border-top-left-radius: 10px;	border-bottom-left-radius: 10px;} QPushButton::hover{background-color:#1f232a;}QPushButton::pressed{background-color:#343b47;}"
            self.ui.AlertasBtn.setStyleSheet(style)
            self.animacion_Vertical_Fin(0,200,500,self.ui.Notificacion)
            self.Side_menu_AL=0
        self.errores()
    def cerrar_MenuN(self):
        style="QPushButton{text-align:left;	padding:5px 10px;	border-top-left-radius: 10px;	border-bottom-left-radius: 10px;} QPushButton::hover{background-color:#1f232a;}QPushButton::pressed{background-color:#343b47;}"
        self.ui.AlertasBtn.setStyleSheet(style)
        self.animacion_Vertical_Fin(0,200,500,self.ui.Notificacion)
        self.Side_menu_AL=0
    def Velocidad(self):
        self.BorrarFormatos()
        self.formato(self.ui.VelocidadBtn)
        self.ui.Contenidos.setCurrentWidget(self.ui.pVel)  
        self.DBPruebas()
        self.ui.TablaVelocidad.resizeRowsToContents()
        for i in range(0,self.ui.TablaVelocidad.columnCount()): self.ui.TablaVelocidad.setColumnWidth(i,self.ui.TablaVelocidad.width()/self.ui.TablaVelocidad.columnCount())
    def DashBoard(self):
        self.BorrarFormatos()
        self.formato(self.ui.DashboardBtn)
        self.ui.Contenidos.setCurrentWidget(self.ui.pTab)
        self.actualizarDatosDash() 
        self.crearTablas()
    senalS=Signal(int)
    def senal(self):
        while self.inf.SALIR!=True:
            self.senalS.emit(1)
            sleep(5)
    def PromedioVel(self):
        pr=db.consultarPRUEBAS_VELOCIDAD()
        datos=[]
        fechav=""
        cont=0
        for i in pr:
            if self.inf.SSID==i[5]:
                if fechav!=i[1]:
                    datos.append([i[1],i[6],i[7]])
                    cont+=1
                else:
                    datos[cont-1][1]=(datos[cont-1][1]+i[6])/2
                    datos[cont-1][2]=(datos[cont-1][2]+i[7])/2
                fechav=i[1]
        return datos
    def crearTablas(self):
        self.crearGraficaVel()
        self.crearGraficaUser()
        self.pendientes()
        disp = QVBoxLayout()
        disp.addWidget(QLabel(f"<p align='center'><b>Buscando Dispositivos en {self.inf.SSID}...</b></p>")) 
        aux2=QWidget()
        aux2.setLayout(disp)
        aux2.setContentsMargins(0,0,0,0)
        self.ui.scrollArea_2.setWidget(aux2)
    def pendientes(self):
        disp = QVBoxLayout()
        disp.addWidget(QLabel("<p align='center'><b>Pendientes</b></p>"))
        pr=db.consultarPROBLEMAS_RED()
        enc=False
        for NE in pr:
            if NE[4]!="Corregido":
                aux=pendientesClass(NE,self.inf.TEMA,self.inf)
                disp.addWidget(aux)
                enc=True
        if enc==False:
            aux=QFrame(self)
            Layout = QGridLayout()
            label = QLabel(self)
            pixmap = QPixmap('Imagenes/CORRECTO.png').scaled(50,50)
            label.setPixmap(pixmap)
            Aux2=QWidget()
            Aux = QHBoxLayout()
            Aux2.setLayout(Aux)
            Layout.addWidget(label,0,0)
            Layout.addWidget(QLabel(""),0,1)
            Layout.addWidget(QLabel("Excelente, sin pendientes"),0,2)
            Layout.addWidget(QLabel(""),0,3)
            if self.inf.TEMA=="Obscuro":
                aux.setStyleSheet("QLabel{border-color:black;} QFrame{Background-color:#1f232a; Border-Radius:10px; }")
            else:
                aux.setStyleSheet("QLabel{border-color:gray;} QFrame{Background-color:#D8D8D8; Border-Radius:10px; }")
            aux.setLayout(Layout)
            disp.addWidget(aux)
        aux2=QWidget()
        aux2.setLayout(disp)
        aux2.setContentsMargins(0,0,0,0)
        self.ui.WidgetPend.setWidget(aux2)
    def crearGraficaUser(self):
        if self.inf.TEMA=="Obscuro":
            enc=["<font color='white'>","</font>"]
        else:
            enc=["<font color='black'>","</font>"]
        series = QBarSeries()
        d=db.consultar("SELECT R.ID_RED, R.SSID,C.MAC_HOST FROM CONTROL_RED_ESPECIFICA AS C INNER JOIN RED AS R ON R.ID_RED=C.ID_RED ORDER BY R.ID_RED")
        datos=[]
        redv=""
        cont=0
        for i in d:
            if redv!=i[1]:
                datos.append([i[1],1])
                cont+=1  
            else:
                datos[cont-1][1]=datos[cont-1][1]+1 
            redv=i[1]
        for i in datos:
            set = QBarSet(i[0])
            set.append(i[1])
            series.append(set)
        series.setLabelsVisible(False)
        cat=QBarCategoryAxis()
        cat.append(["Usuarios"])
        chart =QChart()
        if self.inf.TEMA=="Obscuro":
            chart.setBackgroundVisible(False)
        else:
            chart.setBackgroundVisible(True)
        chart.addSeries(series)
        chart.createDefaultAxes()
        chart.setAxisX(cat,series)
        chart.setAnimationOptions(QChart.SeriesAnimations)
        chart.setTitle(enc[0]+"Numero de usuarios en redes"+enc[1])
        chart.legend().setVisible(True)
        chart.legend().setAlignment(Qt.AlignBottom)
        chartV = QChartView(chart)
        chartV.setRenderHint(QPainter.Antialiasing)
        aux=QWidget()
        layout=QHBoxLayout()
        layout.addWidget(chartV)
        aux.setLayout(layout)
        self.ui.WidgetUser.setWidget(aux)
    def crearGraficaVel(self):
        if self.inf.TEMA=="Obscuro":
            enc=["<font color='white'>","</font>"]
        else:
            enc=["<font color='black'>","</font>"]
        series = QLineSeries()
        series2 = QLineSeries()
        series3= QLineSeries()
        series4 = QLineSeries()
        series5 = QLineSeries()
        aux=self.PromedioVel()
        r=db.consultar(f"SELECT ID_RED,SUBIDA,BAJADA FROM RED WHERE SSID='{self.inf.SSID}'")
        try:
            subida=r[0][1]  
            bajada=r[0][2]
        except:
            subida=0 
            bajada=0
        series.setName(enc[0]+"Subida"+enc[1])
        series2.setName(enc[0]+"Bajada"+enc[1])
        series4.setName(enc[0]+f"{subida}"+enc[1])
        series4.setColor("red")
        series5.setName(enc[0]+f"{bajada}"+enc[1])
        series5.setColor("red")
        series3.hide()
        if self.inf.CONEXION==True:
            dias=0
            if len(aux)<7:
                for i in aux:
                    dias+=1
                    series3.append(dias,0)
                    series3.append(dias,subida+10)
                    series3.append(dias,bajada+10)
                    series.append(dias,i[1])
                    series2.append(dias,i[2])
                    series4.append(dias,subida)
                    series5.append(dias,bajada)
            else:
                dias2=0
                for i in aux:
                    if dias>len(aux)-7:
                        dias2+=1
                        series3.append(dias,0)
                        series3.append(dias,subida+10)
                        series3.append(dias,bajada+10)
                        series.append(dias2,i[1])
                        series2.append(dias2,i[2])
                        series4.append(dias2,subida)
                        series5.append(dias2,bajada)
                    dias+=1
        else:
            series.append(0,0)
            series2.append(0,0)
            series3.append(0,0)
            series4.append(0,0)
            series5.append(0,0)
        chart =QChart()
        if self.inf.TEMA=="Obscuro":
            chart.setBackgroundVisible(False)
        else:
            chart.setBackgroundVisible(True)
        chart.addSeries(series)
        chart.addSeries(series2)
        chart.addSeries(series3)
        chart.addSeries(series4)
        chart.addSeries(series5)
        chart.createDefaultAxes()
        chart.setAnimationOptions(QChart.SeriesAnimations)
        chart.setTitle(enc[0]+"Registro de velocidad de los ultimos 7 Días"+enc[1])
        chart.legend().setVisible(True)
        chart.legend().setAlignment(Qt.AlignBottom)

        chartV = QChartView(chart)
        chartV.setRenderHint(QPainter.Antialiasing)
        aux=QWidget()
        layout=QHBoxLayout()
        layout.addWidget(chartV)
        aux.setLayout(layout)
        self.ui.WidgetVeloci.setWidget(aux)
    def crearGraficaSignal(self):
        try:
            try: senal=float(rs.obtenerIntencidad(self.inf.INTERFAZ).replace("%",""))
            except:senal=0
            pie=QPieSeries()
            if self.inf.CONEXION==True:
                if "ETHERNET" in self.inf.INTERFAZ.upper() and "ETHER" in self.inf.INTERFAZ.upper():
                    senal=100
                    pie.append("Señal (Alámbrico)",senal)
                    pie.append("Perdida",0)
                else:   
                    pie.append("Señal",senal)
                    pie.append("Perdida",100-senal)
            else:
                pie.append("Señal",0)
                pie.append("Perdida",100)
            chart =QChart()
            if self.inf.TEMA=="Obscuro":
                enc=["<font color='white'>","</font>"]
                chart.setBackgroundVisible(False)
            else:
                enc=["<font color='black'>","</font>"]
                chart.setBackgroundVisible(True)
            chart.createDefaultAxes()
            chart.setTitle(enc[0]+"Señal de la red "+self.inf.SSID+f" ({senal}%)"+enc[1])
            chart.addSeries(pie)
            chartV = QChartView(chart)
            chartV.setRenderHint(QPainter.Antialiasing)
            aux=QWidget()
            layout=QHBoxLayout()
            layout.addWidget(chartV)
            aux.setLayout(layout)
            self.ui.WidgetSenal.setWidget(aux)
        except:
            pass

    def actualizarDatosDash(self):
        INTER=rs.obtenerDatosInter(self.inf.INTERFAZ)
        
        self.ui.DatosRed.setText(
            f"""
            <p align='center'> Datos de Red Actual<br> ___________________________________________</p>
            <p align='justify'>
            <b>Nombre Interfaz:</b> {self.inf.INTERFAZ}<br>
            <b>Descripción:</b><br>&nbsp;&nbsp;{INTER[1]}<br>
            <b>SSID:</b> {self.inf.SSID}<br>
            <b>Tipo de radio:</b> {INTER[3]}<br>
            <b>Banda:</b> {INTER[4]}<br>
            <b>Velocidad de Transmision:</b> {INTER[5]} Mbps<br>
            <b>Velocidad de Recepción:</b> {INTER[6]} Mbps<br>
            <b>Señal:</b> {INTER[7]}<br>
            <b>Dirección IPv4:</b> {self.inf.getIPV4()}<br>
            <b>Puerta de Enlace:</b> {self.inf.GATEWAY}<br>
            <b>Sub-Máscara de red:  {self.inf.MASCARA}</b><br>
            </p>
            """
            )

    def Gestion(self):
        self.BorrarFormatos()
        self.formato(self.ui.GestionBtn)
        self.ui.Contenidos.setCurrentWidget(self.ui.pGest) 
        self.ui.SalidaGestion.resizeRowsToContents()
        for i in range(0,self.ui.SalidaGestion.columnCount()):
            self.ui.SalidaGestion.setColumnWidth(i,self.ui.SalidaGestion.width()/self.ui.SalidaGestion.columnCount())
    def Reportes(self):
        self.BorrarFormatos()
        self.formato(self.ui.RerportesBtn)
        self.ui.Contenidos.setCurrentWidget(self.ui.pReport)
    def Trafico(self):
        self.BorrarFormatos()
        self.formato(self.ui.TraficoBtn)
        self.ui.Contenidos.setCurrentWidget(self.ui.pTraf)
        self.ui.TablaPaquetes.resizeRowsToContents()
        for i in range(0,self.ui.TablaPaquetes.columnCount()): self.ui.TablaPaquetes.setColumnWidth(i,self.ui.TablaPaquetes.width()/self.ui.TablaPaquetes.columnCount())
    def Utilidades(self):
        self.BorrarFormatos()
        self.formato(self.ui.UtilidadesBtn)
        self.ui.Contenidos.setCurrentWidget(self.ui.pUtil) 
    
    def IPCONFIG(self):
        self.BorrarFormatos2()
        self.formato(self.ui.IPConfBtn)
        self.ui.ContenidosU.setCurrentWidget(self.ui.IPConfigPag) 
    def PING(self):
        self.BorrarFormatos2()
        self.formato(self.ui.PingBtn)
        self.ui.ContenidosU.setCurrentWidget(self.ui.PingPag) 
        self.ui.PINGOUT.resizeRowsToContents()
        for i in range(0,self.ui.PINGOUT.columnCount()): self.ui.PINGOUT.setColumnWidth(i,self.ui.PINGOUT.width()/self.ui.PINGOUT.columnCount())
    def TRACERT(self):
        self.BorrarFormatos2()
        self.formato(self.ui.TracerBtn)
        self.ui.ContenidosU.setCurrentWidget(self.ui.TracertPag) 
    def ARP(self):
        self.BorrarFormatos2()
        self.formato(self.ui.ArpBtn)
        dis=rs.Obtener_Dispositivos(True)
        self.ui.BarDis.clear()
        for i in dis:
            if i[5] != '--' and i[3]!='127.0.0.1':
                self.ui.BarDis.addItem(i[0]+":"+i[3])
        self.ui.ContenidosU.setCurrentWidget(self.ui.ARPPag) 
    def ROUTE(self):
        self.BorrarFormatos2()
        self.formato(self.ui.RouteBtn)
        self.ui.ContenidosU.setCurrentWidget(self.ui.RoutePag) 
    def NS(self):
        self.BorrarFormatos2()
        self.formato(self.ui.NSLookBtn)
        self.ui.ContenidosU.setCurrentWidget(self.ui.NSPag) 
    def BorrarFormatos(self):
        u=self.ui
        if self.inf.TEMA=="Obscuro":
            style="QPushButton{text-align:left;	padding:5px 10px;	border-top-left-radius: 10px;	border-bottom-left-radius: 10px;} QPushButton::hover{background-color:#1f232a;}QPushButton::pressed{background-color:#343b47;}"
        else:
            style="QPushButton{text-align:left;	padding:5px 10px;	border-top-left-radius: 10px;	border-bottom-left-radius: 10px;} "
        u.DashboardBtn.setStyleSheet(style)
        u.GestionBtn.setStyleSheet(style)
        u.RerportesBtn.setStyleSheet(style)
        u.TraficoBtn.setStyleSheet(style)
        u.UtilidadesBtn.setStyleSheet(style)
        u.VelocidadBtn.setStyleSheet(style)
    def BorrarFormatos2(self):
        u=self.ui
        if self.inf.TEMA=="Obscuro":
            style="QPushButton:hover{background-color:#1f232a;}QPushButton:pressed{	background-color:#343b47;}"
        else:
            style="QPushButton:hover{background-color:#909090;}"
        u.IPConfBtn.setStyleSheet(style)
        u.PingBtn.setStyleSheet(style)
        u.ArpBtn.setStyleSheet(style)
        u.TracerBtn.setStyleSheet(style)
        u.RouteBtn.setStyleSheet(style)
        u.IPConfBtn.setStyleSheet(style)
        u.NSLookBtn.setStyleSheet(style)
    def formato(self,boton):
        if self.inf.TEMA=="Obscuro":
            style="QPushButton{background-color:#1f132a;} QPushButton::hover{background-color:#1f232a;}QPushButton::pressed{background-color:#343b47;}"
        else:
            style="border: 1px solid;border-color:#D8D8D8;"

        boton.setStyleSheet(style)
    ping=""
    procPing=False
    def utilidadPingIniciar(self):
        if self.procPing==False:
            self.procPing=True
            self.ui.PingIniciarBtn.setStyleSheet("border-radius:10px;background-color: #20945e;	padding: 8px 8px;")
            self.ping=""
            busqueda = Thread(target=self.utilidadPingIniciarH)
            busqueda.start()
        else:
            self.procPing=False
    registros=0
    tmpBytes=""

    def putPing(self):
        if "[!]" in self.ping:
            self.ping=self.ping.replace("\n","<br>")
            self.ui.label_35.setText(f"<font color='red'>{self.ping}</font>")
        else:
            valores=self.ping.split(" ")
            self.ui.label_35.setText(f"<font color='cyan'>Haciendo ping a {valores[1]}</font>")
            valores[6]=valores[6].replace("tiempo<","tiempo=<")
            self.ui.PINGOUT.insertRow(self.registros)
            self.ui.PINGOUT.setItem(self.registros,0,QTableWidgetItem(valores[0]))
            self.ui.PINGOUT.setItem(self.registros,1,QTableWidgetItem(valores[1]))
            if "inaccesible" in self.ping:
                self.ui.PINGOUT.setItem(self.registros,3,QTableWidgetItem("Inaccesible"))
                self.ui.PINGOUT.setItem(self.registros,4,QTableWidgetItem("Inaccesible"))
                self.ui.PINGOUT.setItem(self.registros,2,QTableWidgetItem("Inaccesible"))
            elif "agotado" in self.ping:
                self.ui.PINGOUT.setItem(self.registros,2,QTableWidgetItem("Agotado"))
                self.ui.PINGOUT.setItem(self.registros,3,QTableWidgetItem("Agotado"))
                self.ui.PINGOUT.setItem(self.registros,4,QTableWidgetItem("Agotado"))
            else:
                if "bytes=" in self.ping:
                    self.ui.PINGOUT.setItem(self.registros,2,QTableWidgetItem(valores[5].replace("bytes=","")))
                    self.ui.PINGOUT.setItem(self.registros,3,QTableWidgetItem(valores[6].replace("tiempo=","")))
                    self.ui.PINGOUT.setItem(self.registros,4,QTableWidgetItem(valores[7].replace("TTL=","")))
                else:
                    self.ui.PINGOUT.setItem(self.registros,2,QTableWidgetItem(self.tmpBytes))
                    self.ui.PINGOUT.setItem(self.registros,3,QTableWidgetItem(valores[5].replace("tiempo=","")))
                    self.ui.PINGOUT.setItem(self.registros,4,QTableWidgetItem("64"))
            self.registros+=1
    def utilidadPingIniciarH(self):
        print("UTILIDAD: INICIAR PING")
        self.registros=0
        self.tmpBytes=self.ui.ping3.text()
        self.ui.PINGOUT.clearContents()
        a= self.ui.PINGOUT.rowCount()
        for i in range(0,a):   self.ui.PINGOUT.removeRow(i)
        util.utilidad_Ping(self.ui.ping1.toPlainText(),self.ui.ping12.text(),self.ui.ping2.text(),self.tmpBytes,self.ui.ping4.text(),self)
        if self.inf.TEMA=="Obscuro":
            self.ui.PingIniciarBtn.setStyleSheet("border-radius:10px;background-color: #16191d;	padding: 8px 8px;")
        else:
            self.ui.PingIniciarBtn.setStyleSheet("border-radius:10px;background-color: #D8D8D8;	padding: 8px 8px;")

    def utilidadPingAyuda(self):
        var=util.help_ping()
        var=var.replace("Opci¢n incorrecta --help.","")
        var=var.replace("\n","<br>")
        var=var.replace("\t","&nbsp;&nbsp;&nbsp;&nbsp;")
        var=var.replace("ú","u")
        if self.inf.TEMA=="Obscuro":
            self.ui.label_35.setText(f"<font color='white' align='justify'>{var}</font>")
        else:
            self.ui.label_35.setText(f"<font color='black' align='justify'>{var}</font>")
    def IPCMANUAL(self):
        util.IPCMANUAL(self.ui.ParamIp.toPlainText(),self.ui.SalidaIPConf)
    def IPCAll(self):
        util.IPCAll(self.ui.SalidaIPConf)
    def IPCDDNS(self):
        util.IPCDDNS(self.ui.SalidaIPConf)
    def IPCFDNS(self):
        util.IPCFDNS(self.ui.SalidaIPConf)
    def IPCRDNS(self):
        util.IPCRDNS(self.ui.SalidaIPConf)
    def IPCAYUDA(self):
        util.IPCAYUDA(self.ui.SalidaIPConf)
    def ARPA(self):
        util.ARPA(self.ui.SalidaARP)
    def ARPC(self):
        util.ARPC(self.ui.arp6.text(), self.ui.SalidaARP,self.inf.INTERFAZ)
    def ARPADD(self):
        util.ARPADD(self.ui.arp4.text(),self.ui.arp3.text(),self.ui.BarDis.itemText(self.ui.BarDis.currentIndex()).split(":")[1],self.ui.SalidaARP)
    def ARPAY(self):
        util.ARPAY(self.ui.SalidaARP)
    def ARPR(self):
        util.ARPR(self.ui.arp8.text(),self.ui.BarDis.itemText(self.ui.BarDis.currentIndex()).split(":")[1],self.ui.SalidaARP)
    def NSLOI(self):
        util.NSLOI(self.ui.nslo1.text(),self.ui.nslo2.text(),self.ui.textEdit)
    def NSLOA(self):
        util.NSLOA(self.ui.textEdit)
    def tracertA(self):
        util.tracertA(self.ui.SalidaTracert)
    def tracertI(self):
        util.tracertI(self.ui.tracert0.text(),self.ui.tracert1.text(),self.ui.tracert2.text(),self.ui.tracert3.text(),self.ui.tracert4.isChecked(),self.ui.tracert5.isChecked(),self.ui.SalidaTracert)
    procesoT=False
    def trafico(self):
        if self.TrafObj.Cerrar==False:
            a= self.ui.TablaPaquetes.rowCount()
            self.paquetes=1
            for i in range(1,a):   self.ui.TablaPaquetes.removeRow(i)
            self.ui.TablaPaquetes.clearContents()
            self.TrafObj.Cerrar=True
            self.ui.InicioTraficoBtn.setStyleSheet("border-radius:10px;background-color: #20945e;	padding: 8px 8px;")
            hilotrafico=Thread(target=self.captura)
            hilotrafico.start()
        else:
            self.TrafObj.Cerrar=False
            if self.inf.TEMA=="Obscuro":
                self.ui.InicioTraficoBtn.setStyleSheet("border-radius:10px;background-color: #16191d;	padding: 8px 8px;")
            else:
                self.ui.InicioTraficoBtn.setStyleSheet("background-color:#D8D8D8;border-radius:10px;padding: 8px 8px;")
    def captura(self):
        rs.monitor(self.inf.INTERFAZ,self.signalTrafico,self.TrafObj,self.inf)
    paquetes=1
    SalidaPaq=False
    def dominios(self):
        if self.SalidaPaq==False:
            self.ui.TituloSalidaPaq.setText("Dominios capturados")
            self.animacion_Vertical_Inicio(0,200,500,self.ui.SalidaPaq)
            texto=""
            with open('modulos/BasesDatos/Dominios.txt', "r",encoding="utf-8") as archivo_lectura:
                for linea in archivo_lectura:
                    texto+=linea+"<br>"
            self.ui.TextoSalidaPaq.setHtml(texto)
            self.SalidaPaq=True
        else:
            self.SalidaPaq=False
            self.animacion_Vertical_Fin(0,200,500,self.ui.SalidaPaq)
    def salirPaq(self):
        self.SalidaPaq=False
        self.animacion_Vertical_Fin(0,200,500,self.ui.SalidaPaq)
    def Doubleclick(self):
        self.ui.TituloSalidaPaq.setText("Información del Paquete")
        for idx in self.ui.TablaPaquetes.selectionModel().selectedIndexes():
                row_number = idx.row()
        file=open('modulos/BasesDatos/sniff.pcap','r',encoding='utf-8')
        contenido=file.readlines()[int(self.ui.TablaPaquetes.item(row_number,0).text())]
        if self.SalidaPaq==False:
            self.animacion_Vertical_Inicio(0,200,500,self.ui.SalidaPaq)
            self.SalidaPaq=True
            self.ui.TextoSalidaPaq.setHtml(rs.formatoPaquete(contenido))
        else:
            self.ui.TextoSalidaPaq.setHtml(rs.formatoPaquete(contenido))
        file.close()
    FiltrarP=""
    
    def FiltrarPaquete(self):
        self.paquetes=0
        self.FiltrarP=self.ui.Filtrado.text()
        if self.FiltrarP.replace(" ","")=="":
            for i in range(0,self.ui.TablaPaquetes.rowCount()):   self.ui.TablaPaquetes.removeRow(i)
            self.ui.TablaPaquetes.setRowCount(0)
            with open("modulos/BasesDatos/sniff.pcap", "r",encoding="utf-8") as archivo_lectura:
                for linea in archivo_lectura:
                    aux=rs.get_data(str(linea))
                    self.ui.TablaPaquetes.insertRow(self.paquetes)
                    self.ui.TablaPaquetes.setItem(self.paquetes,0,QTableWidgetItem(str(aux[0])))
                    protocolo=QTableWidgetItem(aux[2])
                    protocolo.setForeground(QColor(255, 255, 255))
                    if str(aux[2]) == "TCP": 
                        protocolo.setBackground(QColor(86, 82, 100))
                    elif str(aux[2]) == "UDP": protocolo.setBackground(QColor(83, 153, 176))
                    elif str(aux[2]) == "ICMP": protocolo.setBackground(QColor(21, 52, 80))
                    elif str(aux[2]) == "ARP": protocolo.setBackground(QColor(41, 64, 82))
                    elif str(aux[2]) == "DNS": protocolo.setBackground(QColor(105, 162, 151))
                    else: protocolo.setBackground(QColor(80, 128, 142))

                    self.ui.TablaPaquetes.setItem(self.paquetes,0,QTableWidgetItem(aux[0]))
                    self.ui.TablaPaquetes.setItem(self.paquetes,1,QTableWidgetItem(str(aux[1])))
                    self.ui.TablaPaquetes.setItem(self.paquetes,2,protocolo)
                    self.ui.TablaPaquetes.setItem(self.paquetes,3,QTableWidgetItem(str(aux[3])))
                    self.ui.TablaPaquetes.setItem(self.paquetes,4,QTableWidgetItem(str(aux[4])))
                    self.ui.TablaPaquetes.setItem(self.paquetes,5,QTableWidgetItem(str(aux[5])))
                    self.ui.TablaPaquetes.setItem(self.paquetes,6,QTableWidgetItem(str(aux[6])))
                    self.paquetes+=1
        else:
            for i in range(0,self.ui.TablaPaquetes.rowCount()):   self.ui.TablaPaquetes.removeRow(i)
            self.ui.TablaPaquetes.setRowCount(0)
            with open("modulos/BasesDatos/sniff.pcap", "r",encoding="utf-8") as archivo_lectura:
                for linea in archivo_lectura:
                    if self.FiltrarP.upper() in linea.upper():
                        aux=rs.get_data(str(linea))
                        self.ui.TablaPaquetes.insertRow(self.paquetes)
                        protocolo=QTableWidgetItem(aux[2])
                        protocolo.setForeground(QColor(255, 255, 255))
                        if str(aux[2]) == "TCP": protocolo.setBackground(QColor(86, 82, 100))
                        elif str(aux[2]) == "UDP": protocolo.setBackground(QColor(83, 153, 176))
                        elif str(aux[2]) == "ICMP": protocolo.setBackground(QColor(21, 52, 80))
                        elif str(aux[2]) == "ARP": protocolo.setBackground(QColor(41, 64, 82))
                        elif str(aux[2]) == "DNS": protocolo.setBackground(QColor(105, 162, 151))
                        else: protocolo.setBackground(QColor(80, 128, 142))

                        self.ui.TablaPaquetes.setItem(self.paquetes,0,QTableWidgetItem(str(aux[0])))
                        self.ui.TablaPaquetes.setItem(self.paquetes,1,QTableWidgetItem(str(aux[1])))
                        self.ui.TablaPaquetes.setItem(self.paquetes,2,protocolo)
                        self.ui.TablaPaquetes.setItem(self.paquetes,3,QTableWidgetItem(str(aux[3])))
                        self.ui.TablaPaquetes.setItem(self.paquetes,4,QTableWidgetItem(str(aux[4])))
                        self.ui.TablaPaquetes.setItem(self.paquetes,5,QTableWidgetItem(str(aux[5])))
                        self.ui.TablaPaquetes.setItem(self.paquetes,6,QTableWidgetItem(str(aux[6])))
                        self.paquetes+=1
    def BorrarTabla(self):
        self.TrafObj.borrar=True
        a= self.ui.TablaPaquetes.rowCount()
        self.paquetes=1
        for i in range(1,a):   self.ui.TablaPaquetes.removeRow(i)
        self.ui.TablaPaquetes.clearContents()
        Salida=open('modulos/BasesDatos/sniff.pcap',"w",encoding="utf-8")
        Salida.close()
    indicador="PR"
    indicadorV=""
    def DBBorrarFormatos(self):
        u=self.ui
        if self.inf.TEMA=="Obscuro":
            style="QPushButton{text-align:center; border-top-left-radius: 10px;} QPushButton::hover{background-color:#1f232a;}QPushButton::pressed{background-color:#343b47;}"
        else:
            style="QPushButton{text-align:center; border-top-left-radius: 10px;}QPushButton::hover{background-color:#D8D8D8;}"
        u.G_PR.setStyleSheet(style)
        u.G_R.setStyleSheet(style)
        u.G_I.setStyleSheet(style)
        u.G_P.setStyleSheet(style)
        u.G_RD.setStyleSheet(style)
        u.G_C.setStyleSheet(style)
        u.G_H.setStyleSheet(style)
        u.G_CR.setStyleSheet(style)
        u.G_SU.setStyleSheet(style)
        u.G_D.setStyleSheet(style)
    def seccionadoDB(self,boton,ind):
        self.DBBorrarFormatos()
        self.formato(boton)
        self.indicador=ind
        self.DBcerrarMenuAct()
    Filtrado="*_*_*"
    Filtrado2="*_*_*"
    def FiltrarSpeed(self):
        aux=self.ui.FiltradoSpeed.text()
        if aux==" " or aux=="\t":
            self.Filtrado2="*_*_*"
        else:
            self.Filtrado2=aux
        self.DBPruebas()
    def agregarF(self,tabla,aux,num):
        reg=0
        if self.Filtrado=="*_*_*":
            for i in aux:
                tabla.insertRow(reg)
                for a in range(0,num):
                    if "BIEN" in str(i[a]).upper():
                        formato=QTableWidgetItem(str(i[a]))
                        formato.setForeground(QColor(255, 255, 255))
                        formato.setBackground(QColor(114, 110, 255))
                        tabla.setItem(reg,a,formato)
                    elif "REGULAR" in str(i[a]).upper() or "NO VERIFICADO" in str(i[a]).upper():
                        formato=QTableWidgetItem(str(i[a]))
                        formato.setForeground(QColor(255, 255, 255))
                        formato.setBackground(QColor(255, 103, 50))
                        tabla.setItem(reg,a,formato)
                    elif "MAL" in str(i[a]).upper() or "NO CORREGIDO" in str(i[a]).upper() or "NO PERMITIDO" in str(i[a]).upper():
                        formato=QTableWidgetItem(str(i[a]))
                        formato.setForeground(QColor(255, 255, 255))
                        formato.setBackground(QColor(255, 79, 97))
                        tabla.setItem(reg,a,formato)
                    elif "CORRECTO" in str(i[a]).upper() or "CORREGIDO" in str(i[a]).upper() or "LIBRE" in str(i[a]).upper() or "EXCELENTE" in str(i[a]).upper() or "VERIFICADO" in str(i[a]).upper():
                        formato=QTableWidgetItem(str(i[a]))
                        formato.setForeground(QColor(255, 255, 255))
                        formato.setBackground(QColor(67, 144, 67))
                        tabla.setItem(reg,a,formato)
                    else:
                        tabla.setItem(reg,a,QTableWidgetItem(str(i[a])))
                reg+=1
        else:
            for i in aux:
                if self.filtradoV(self.Filtrado.upper(),str(i).upper())==True:
                    tabla.insertRow(reg)
                    for a in range(0,num):
                        if "BIEN" in str(i[a]).upper():
                            formato=QTableWidgetItem(str(i[a]))
                            formato.setForeground(QColor(255, 255, 255))
                            formato.setBackground(QColor(114, 110, 255))
                            tabla.setItem(reg,a,formato)
                        elif "REGULAR" in str(i[a]).upper() or "NO VERIFICADO" in str(i[a]).upper():
                            formato=QTableWidgetItem(str(i[a]))
                            formato.setForeground(QColor(255, 255, 255))
                            formato.setBackground(QColor(255, 103, 50))
                            tabla.setItem(reg,a,formato)
                        elif "MAL" in str(i[a]).upper() or "NO CORREGIDO" in str(i[a]).upper() or "NO PERMITIDO" in str(i[a]).upper():
                            formato=QTableWidgetItem(str(i[a]))
                            formato.setForeground(QColor(255, 255, 255))
                            formato.setBackground(QColor(255, 79, 97))
                            tabla.setItem(reg,a,formato)
                        elif "CORRECTO" in str(i[a]).upper() or "CORREGIDO" in str(i[a]).upper() or "LIBRE" in str(i[a]).upper() or "EXCELENTE" in str(i[a]).upper() or "VERIFICADO" in str(i[a]).upper():
                            formato=QTableWidgetItem(str(i[a]))
                            formato.setForeground(QColor(255, 255, 255))
                            formato.setBackground(QColor(67, 144, 67))
                            tabla.setItem(reg,a,formato)
                        else:
                            tabla.setItem(reg,a,QTableWidgetItem(str(i[a])))
                    reg+=1
        tabla.resizeRowsToContents()
        for i in range(0,tabla.columnCount()):
            tabla.setColumnWidth(i,tabla.width()/tabla.columnCount())

    def agregarFS(self,tabla,aux,num):
        reg=0
        if self.Filtrado2=="*_*_*":
            for i in aux:
                tabla.insertRow(reg)
                for a in range(0,num):
                    if "BIEN" in str(i[a]).upper() or "BUENO" in str(i[a]).upper():
                        formato=QTableWidgetItem(str(i[a]))
                        formato.setForeground(QColor(255, 255, 255))
                        formato.setBackground(QColor(114, 110, 255))
                        tabla.setItem(reg,a,formato)
                    elif "REGULAR" in str(i[a]).upper() or "NO VERIFICADO" in str(i[a]).upper():
                        formato=QTableWidgetItem(str(i[a]))
                        formato.setForeground(QColor(255, 255, 255))
                        formato.setBackground(QColor(255, 103, 50))
                        tabla.setItem(reg,a,formato)
                    elif "MAL" in str(i[a]).upper() or "NO CORREGIDO" in str(i[a]).upper() or "NO PERMITIDO" in str(i[a]).upper():
                        formato=QTableWidgetItem(str(i[a]))
                        formato.setForeground(QColor(255, 255, 255))
                        formato.setBackground(QColor(255, 79, 97))
                        tabla.setItem(reg,a,formato)
                    elif "CORRECTO" in str(i[a]).upper() or "CORREGIDO" in str(i[a]).upper() or "LIBRE" in str(i[a]).upper() or "EXCELENTE" in str(i[a]).upper() or "VERIFICADO" in str(i[a]).upper():
                        formato=QTableWidgetItem(str(i[a]))
                        formato.setBackground(QColor(67, 144, 67))
                        formato.setForeground(QColor(255, 255, 255))
                        tabla.setItem(reg,a,formato)
                    else:
                        tabla.setItem(reg,a,QTableWidgetItem(str(i[a])))
                reg+=1
        else:
            for i in aux:
                if self.filtradoV(self.Filtrado2.upper(),str(i).upper())==True:
                    tabla.insertRow(reg)
                    for a in range(0,num):
                        if "BIEN" in str(i[a]).upper() or "BUENO" in str(i[a]).upper():
                            formato=QTableWidgetItem(str(i[a]))
                            formato.setForeground(QColor(255, 255, 255))
                            formato.setBackground(QColor(114, 110, 255))
                            tabla.setItem(reg,a,formato)
                        elif "REGULAR" in str(i[a]).upper() or  "NO VERIFICADO" in str(i[a]).upper():
                            formato=QTableWidgetItem(str(i[a]))
                            formato.setForeground(QColor(255, 255, 255))
                            formato.setBackground(QColor(255, 103, 50))
                            tabla.setItem(reg,a,formato)
                        elif "MAL" in str(i[a]).upper() or "NO CORREGIDO" in str(i[a]).upper() or "NO PERMITIDO" in str(i[a]).upper():
                            formato=QTableWidgetItem(str(i[a]))
                            formato.setForeground(QColor(255, 255, 255))
                            formato.setBackground(QColor(255, 79, 97))
                            tabla.setItem(reg,a,formato)
                        elif "CORRECTO" in str(i[a]).upper() or "CORREGIDO" in str(i[a]).upper() or "LIBRE" in str(i[a]).upper() or "EXCELENTE" in str(i[a]).upper() or  "VERIFICADO" in str(i[a]).upper():
                            formato=QTableWidgetItem(str(i[a]))
                            formato.setBackground(QColor(67, 144, 67))
                            formato.setForeground(QColor(255, 255, 255))
                            tabla.setItem(reg,a,formato)
                        else:
                            tabla.setItem(reg,a,QTableWidgetItem(str(i[a])))
                    reg+=1
        tabla.resizeRowsToContents()
        for i in range(0,tabla.columnCount()):
            tabla.setColumnWidth(i,tabla.width()/tabla.columnCount())
    def filtradoV(self,filtrado,valor):
        if "|" in filtrado:
            aux=filtrado.split("|")
            for i in aux:
                if i in valor:
                    return True
        if "&" in filtrado:
            aux=filtrado.split("&")
            encontrado=0
            for i in aux:
                if i in valor:
                    encontrado+=1
            if len(aux)==encontrado:
                return True
        if "|" not in filtrado and "&" not in filtrado:
            if filtrado in valor:
                return True
        return False
    def DBproblemasRed(self):
        self.seccionadoDB(self.ui.G_PR,"PR")
        dep=db.consultarDEPARTAMENTOS()
        self.ui.PR_DEP.clear()
        for i in dep: self.ui.PR_DEP.addItem(str(i[0])+":"+i[1])
        dep=db.consultarRED()
        self.ui.PR_RED.clear()
        for i in dep: self.ui.PR_RED.addItem(str(i[0])+":"+i[1])
        self.borrarTabla()
        for i in range(0,8): self.ui.SalidaGestion.insertColumn(i)
        self.ui.SalidaGestion.setHorizontalHeaderItem(0,QTableWidgetItem("ID"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(1,QTableWidgetItem("Descripción"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(2,QTableWidgetItem("Fecha"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(3,QTableWidgetItem("Solución"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(4,QTableWidgetItem("Corregido"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(5,QTableWidgetItem("Fecha correxión"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(6,QTableWidgetItem("Red"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(7,QTableWidgetItem("Departamento"))
        aux = db.consultarPROBLEMAS_RED()
        self.agregarF(self.ui.SalidaGestion,aux,8)
        
    def DBreportes(self):
        self.seccionadoDB(self.ui.G_R,"R")
        self.borrarTabla()
        dep=db.consultarPROVEEDOR()
        self.ui.R_PROV.clear()
        for i in dep: self.ui.R_PROV.addItem(str(i[0])+":"+i[1])
        for i in range(0,10):  self.ui.SalidaGestion.insertColumn(i)
        self.ui.SalidaGestion.setHorizontalHeaderItem(0,QTableWidgetItem("ID"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(1,QTableWidgetItem("Folio"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(2,QTableWidgetItem("Descripción"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(3,QTableWidgetItem("Fecha"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(4,QTableWidgetItem("Acudio"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(5,QTableWidgetItem("Remoto / Sitio"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(6,QTableWidgetItem("Motivo"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(7,QTableWidgetItem("Telefono"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(8,QTableWidgetItem("Atendio"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(9,QTableWidgetItem("Proveedor"))
        aux = db.consultarREPORTES()
        self.agregarF(self.ui.SalidaGestion,aux,10)


    def DBinvernario(self):
        self.seccionadoDB(self.ui.G_I,"I")
        dep=db.consultarDEPARTAMENTOS()
        self.ui.I_DEP.clear()
        for i in dep: self.ui.I_DEP.addItem(str(i[0])+":"+i[1])
        dep=db.consultarRED()
        self.ui.I_RED.clear()
        for i in dep: self.ui.I_RED.addItem(str(i[0])+":"+i[1])
        self.borrarTabla()
        for i in range(0,8):  self.ui.SalidaGestion.insertColumn(i)
        self.ui.SalidaGestion.setHorizontalHeaderItem(0,QTableWidgetItem("ID"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(1,QTableWidgetItem("No. Inventario"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(2,QTableWidgetItem("Nombre"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(3,QTableWidgetItem("Descripción"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(4,QTableWidgetItem("Ubicación"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(5,QTableWidgetItem("Tipo Conexión"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(6,QTableWidgetItem("Red"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(7,QTableWidgetItem("Departamento"))
        aux = db.consultarDISPOSITIVOS()
        self.agregarF(self.ui.SalidaGestion,aux,8)
    def DBProveedores(self):
        self.seccionadoDB(self.ui.G_P,"P")
        dep=db.consultarTRANSMISION()
        self.ui.P_TR.clear()
        for i in dep: self.ui.P_TR.addItem(str(i[0])+":"+i[1])
        self.borrarTabla()
        for i in range(0,6):  self.ui.SalidaGestion.insertColumn(i)
        self.ui.SalidaGestion.setHorizontalHeaderItem(0,QTableWidgetItem("ID"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(1,QTableWidgetItem("Nombre"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(2,QTableWidgetItem("Descripción"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(3,QTableWidgetItem("Fecha"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(4,QTableWidgetItem("Telefono"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(5,QTableWidgetItem("Tipo"))
        aux = db.consultarPROVEEDOR()
        self.agregarF(self.ui.SalidaGestion,aux,6)
    def DBRedes(self):
        self.seccionadoDB(self.ui.G_RD,"RD")
        dep=db.consultarPROVEEDOR()
        self.ui.RD_PROV.clear()
        for i in dep: self.ui.RD_PROV.addItem(str(i[0])+":"+i[1])
        self.borrarTabla()
        for i in range(0,6):  self.ui.SalidaGestion.insertColumn(i)
        self.ui.SalidaGestion.setHorizontalHeaderItem(0,QTableWidgetItem("ID"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(1,QTableWidgetItem("SSID"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(2,QTableWidgetItem("Contraseña"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(3,QTableWidgetItem("Subida"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(4,QTableWidgetItem("Bajada"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(5,QTableWidgetItem("Proveedor"))
        aux = db.consultarRED()
        self.agregarF(self.ui.SalidaGestion,aux,6)
    def DBControlRed(self):
        self.seccionadoDB(self.ui.G_CR,"CR")
        dep=db.consultarDEPARTAMENTOS()
        self.ui.CR_DEP.clear()
        for i in dep: self.ui.CR_DEP.addItem(str(i[0])+":"+i[1])
        dep=db.consultarRED()
        self.ui.CR_RED.clear()
        for i in dep: self.ui.CR_RED.addItem(str(i[0])+":"+i[1])
        self.borrarTabla()
        for i in range(0,14):  self.ui.SalidaGestion.insertColumn(i)
        self.ui.SalidaGestion.setHorizontalHeaderItem(0,QTableWidgetItem("ID"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(1,QTableWidgetItem("Fecha"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(2,QTableWidgetItem("SSID"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(3,QTableWidgetItem("No. Inventario"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(4,QTableWidgetItem("Estado"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(5,QTableWidgetItem("Host"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(6,QTableWidgetItem("IPv4"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(7,QTableWidgetItem("Vendor"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(8,QTableWidgetItem("MAC"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(9,QTableWidgetItem("Usuario"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(10,QTableWidgetItem("Observaciones"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(11,QTableWidgetItem("Verificado"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(12,QTableWidgetItem("Tipo"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(13,QTableWidgetItem("Departamento"))
        aux = db.consultarCONTROL_RED_ESPECIFICA()
        reg=0
        self.agregarF(self.ui.SalidaGestion,aux,14)

    def DBUnidades(self):
        self.seccionadoDB(self.ui.G_SU,"U")
        dep=db.consultarTRANSMISION()
        self.ui.SU_TR.clear()
        for i in dep: self.ui.SU_TR.addItem(str(i[0])+":"+i[1])
        self.borrarTabla()
        for i in range(0,9):  self.ui.SalidaGestion.insertColumn(i)
        self.ui.SalidaGestion.setHorizontalHeaderItem(0,QTableWidgetItem("ID"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(1,QTableWidgetItem("Coordinación"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(2,QTableWidgetItem("Nombre"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(3,QTableWidgetItem("Con Internet"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(4,QTableWidgetItem("Proveedor"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(5,QTableWidgetItem("Observaciones"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(6,QTableWidgetItem("Fecha"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(7,QTableWidgetItem("Telefono"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(8,QTableWidgetItem("Tipo Transmisión"))
        aux = db.consultarUNIDADES_INTERNET()
        self.agregarF(self.ui.SalidaGestion,aux,9)
        
    def DBDepartamentos(self):
        self.seccionadoDB(self.ui.G_D,"D")
        self.borrarTabla()
        for i in range(0,3):  self.ui.SalidaGestion.insertColumn(i)
        self.ui.SalidaGestion.setHorizontalHeaderItem(0,QTableWidgetItem("ID"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(1,QTableWidgetItem("Nombre"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(2,QTableWidgetItem("Descripcion"))
        aux = db.consultarDEPARTAMENTOS()
        reg=0
        self.agregarF(self.ui.SalidaGestion,aux,3)

    def DBConexiones(self):
        self.seccionadoDB(self.ui.G_C,"C")
        self.borrarTabla()
        for i in range(0,4):  self.ui.SalidaGestion.insertColumn(i)
        self.ui.SalidaGestion.setHorizontalHeaderItem(0,QTableWidgetItem("ID"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(1,QTableWidgetItem("SSID"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(2,QTableWidgetItem("Fecha"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(3,QTableWidgetItem("Conexiones"))
        aux = db.consultarHISTORIAL_NUM_CONEXIONES()
        reg=0
        self.agregarF(self.ui.SalidaGestion,aux,4)
    def DBPruebas(self):
        for i in range(0,self.ui.TablaVelocidad.rowCount()):   self.ui.TablaVelocidad.removeRow(i)
        self.ui.TablaVelocidad.setRowCount(0)
        for i in range(0,self.ui.TablaVelocidad.columnCount()): self.ui.TablaVelocidad.removeColumn(i)
        self.ui.TablaVelocidad.setColumnCount(0)
        for i in range (0,12): self.ui.TablaVelocidad.insertColumn(i)
        self.ui.TablaVelocidad.setHorizontalHeaderItem(0,QTableWidgetItem("ID"))
        self.ui.TablaVelocidad.setHorizontalHeaderItem(1,QTableWidgetItem("Fecha"))
        self.ui.TablaVelocidad.setHorizontalHeaderItem(2,QTableWidgetItem("Hora"))
        self.ui.TablaVelocidad.setHorizontalHeaderItem(3,QTableWidgetItem("Host"))
        self.ui.TablaVelocidad.setHorizontalHeaderItem(4,QTableWidgetItem("Sponsor"))
        self.ui.TablaVelocidad.setHorizontalHeaderItem(5,QTableWidgetItem("Red"))
        self.ui.TablaVelocidad.setHorizontalHeaderItem(6,QTableWidgetItem("Subida"))
        self.ui.TablaVelocidad.setHorizontalHeaderItem(7,QTableWidgetItem("Bajada"))
        self.ui.TablaVelocidad.setHorizontalHeaderItem(8,QTableWidgetItem("Latencia (S)"))
        self.ui.TablaVelocidad.setHorizontalHeaderItem(9,QTableWidgetItem("Latencia (B)"))
        self.ui.TablaVelocidad.setHorizontalHeaderItem(10,QTableWidgetItem("Ping"))
        self.ui.TablaVelocidad.setHorizontalHeaderItem(11,QTableWidgetItem("Resultado"))
        aux = db.consultarPRUEBAS_VELOCIDAD()
        reg=0
        self.agregarFS(self.ui.TablaVelocidad,aux,12)
        

    def DBHosts(self):
        self.seccionadoDB(self.ui.G_H,"H")
        self.borrarTabla()
        for i in range(0,7):  self.ui.SalidaGestion.insertColumn(i)
        self.ui.SalidaGestion.setHorizontalHeaderItem(0,QTableWidgetItem("ID"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(1,QTableWidgetItem("FECHA"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(2,QTableWidgetItem("SSID"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(3,QTableWidgetItem("Host"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(4,QTableWidgetItem("IPv4"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(5,QTableWidgetItem("MAC"))
        self.ui.SalidaGestion.setHorizontalHeaderItem(6,QTableWidgetItem("VENDOR"))
        aux = db.consultarHISTORIAL_CONEXIONES()
        reg=0
        self.agregarF(self.ui.SalidaGestion,aux,7)
    def borrarTabla(self):
        for i in range(0,self.ui.SalidaGestion.rowCount()):   self.ui.SalidaGestion.removeRow(i)
        self.ui.SalidaGestion.setRowCount(0)
        for i in range(0,self.ui.SalidaGestion.columnCount()): self.ui.SalidaGestion.removeColumn(i)
        self.ui.SalidaGestion.setColumnCount(0)
    def DBcerrarMenuAct(self):
        if self.indicadorV=="PR":
            self.animacion_Vertical_Fin(0,200,500,self.ui.P_PR)
        elif self.indicadorV=="R":
            self.animacion_Vertical_Fin(0,200,500,self.ui.P_R)
        elif self.indicadorV=="I":
            self.animacion_Vertical_Fin(0,200,500,self.ui.P_I)
        elif self.indicadorV=="P":
            self.animacion_Vertical_Fin(0,200,500,self.ui.P_P)
        elif self.indicadorV=="RD":
            self.animacion_Vertical_Fin(0,200,500,self.ui.P_RD)
        elif self.indicadorV=="CR":
            self.animacion_Vertical_Fin(0,200,500,self.ui.P_CR)
        elif self.indicadorV=="U":
            self.animacion_Vertical_Fin(0,200,500,self.ui.P_SU)
        elif self.indicadorV=="D":
            self.animacion_Vertical_Fin(0,200,500,self.ui.P_D)
        self.limpiarCampos()
        self.indicadorV=""
    def DBAbrirMenuAct(self):
        if self.indicadorV!=self.indicador:
            self.DBcerrarMenuAct()
            if self.indicador=="PR":
                self.animacion_Vertical_Inicio(0,200,500,self.ui.P_PR)
            elif self.indicador=="R":
                self.animacion_Vertical_Inicio(0,200,500,self.ui.P_R)
            elif self.indicador=="I":
                self.animacion_Vertical_Inicio(0,200,500,self.ui.P_I)
            elif self.indicador=="P":
                self.animacion_Vertical_Inicio(0,200,500,self.ui.P_P)
            elif self.indicador=="RD":
                self.animacion_Vertical_Inicio(0,200,500,self.ui.P_RD)
            elif self.indicador=="CR":
                self.animacion_Vertical_Inicio(0,200,500,self.ui.P_CR)
            elif self.indicador=="U":
                self.animacion_Vertical_Inicio(0,200,500,self.ui.P_SU)
            elif self.indicador=="D":
                self.animacion_Vertical_Inicio(0,200,500,self.ui.P_D)
            self.indicadorV=self.indicador
        else:
            if self.indicadorM==False:
                self.DBcerrarMenuAct()
                formato="border-radius:10px;padding: 8px 8px;"
                self.ui.GModificarBtn.setStyleSheet(formato)
                self.ui.GEliminarBtn.setStyleSheet(formato)
                self.ui.GModificarBtn.setEnabled(False)
                self.ui.GEliminarBtn.setEnabled(False)
                formato="border-radius:10px;background-color: #20945e;	padding: 8px 8px;"
                self.ui.CR_ADD.setStyleSheet(formato)
                self.ui.D_ADD.setStyleSheet(formato)
                self.ui.I_ADD.setStyleSheet(formato)
                self.ui.P_ADD.setStyleSheet(formato)
                self.ui.G_AgregarBtn.setStyleSheet(formato)
                self.ui.R_ADD.setStyleSheet(formato)
                self.ui.RD_ADD.setStyleSheet(formato)
                self.ui.SU_ADD.setStyleSheet(formato)
                self.ui.D_ADD.setEnabled(True)
                self.ui.I_ADD.setEnabled(True)
                self.ui.R_ADD.setEnabled(True)
                self.ui.P_ADD.setEnabled(True)
                self.ui.CR_ADD.setEnabled(True)
                self.ui.RD_ADD.setEnabled(True)
                self.ui.SU_ADD.setEnabled(True)
                self.ui.G_AgregarBtn.setEnabled(True)
                self.ui.SU_ID.setEnabled(True)
    indicadorM=False
    def DBagregar(self):
        self.indicadorM=False
        self.DBAbrirMenuAct()
    def DBCR_ADD(self):
        db.insertarCONTROL_RED_ESPECIFICA(int(self.TextoComboBox(self.ui.CR_RED).split(":")[0]),self.ui.CR_IP.text(),self.ui.CR_NOINV.text(),self.TextoComboBox(self.ui.CR_STAT),self.ui.CR_MAC.text(),self.ui.CR_RESP.text(),self.ui.CR_FECHA.text(),self.ui.CR_OBS.toPlainText(),self.ui.CR_HOS.text(),self.TextoComboBox(self.ui.CR_VERFI),self.TextoComboBox(self.ui.CR_TIP),int(self.TextoComboBox(self.ui.CR_DEP).split(":")[0]),self.ui.CR_VEND.text())
        self.DBControlRed()
    def DBD_ADD(self):
        db.insertarDEPARTAMENTOS(self.ui.D_NOM.text(),self.ui.D_DES.toPlainText())
        self.DBDepartamentos()
    def DBI_ADD(self):
        db.insertarDISPOSITIVOS(self.ui.I_ID.text(),self.ui.I_NOM.text(),self.ui.I_DESC.toPlainText(),self.ui.I_UBI.text(),self.TextoComboBox(self.ui.I_TC),int(self.TextoComboBox(self.ui.I_RED).split(":")[0]),int(self.TextoComboBox(self.ui.I_DEP).split(":")[0]))
        self.DBinvernario()
    def DBP_ADD(self):
        db.insertarPROVEEDOR(self.ui.P_N.text(),self.ui.P_DES.toPlainText(),self.ui.P_FECH.text(),self.ui.P_TEL.text(),int(self.TextoComboBox(self.ui.P_TR).split(":")[0]))
        self.DBProveedores()
    def DBP_PR(self):
        db.insertarPROBLEMAS_RED(self.ui.PR_DESC.toPlainText(),self.ui.PR_FECHA.text(),self.ui.PR_SOL.toPlainText(),self.TextoComboBox(self.ui.PR_COR),self.ui.PR_FECHAC.text(),int(self.TextoComboBox(self.ui.PR_RED).split(":")[0]),int(self.TextoComboBox(self.ui.PR_DEP).split(":")[0]))
        self.DBproblemasRed()
    def DBP_R(self):
        db.insertarREPORTES(self.ui.R_Folio.text(),self.ui.R_Descrip.toPlainText(),self.ui.R_Fecha.text(),self.TextoComboBox(self.ui.R_ACUDIO),self.TextoComboBox(self.ui.R_REMOSIT),self.ui.R_MOT.toPlainText(),self.ui.R_TEL.text(),self.ui.R_ATEND.text(),int(self.TextoComboBox(self.ui.R_PROV).split(":")[0]))
        self.DBreportes()
    def DBP_RD(self):
        db.insertarRED(self.ui.RD_S.text(),self.ui.RD_PASS.text(),float(self.ui.RD_SUB.text()),float(self.ui.RD_DAJ.text()),int(self.TextoComboBox(self.ui.RD_PROV).split(":")[0]))
        self.DBRedes()
    def DBP_SU(self):
        db.insertarUNIDADES_INTERNET(self.ui.SU_ID.text(),self.ui.SU_NOM.text(),self.ui.SU_CORD.text(),self.TextoComboBox(self.ui.SU_SERV),self.ui.SU_PROVE.text(),self.ui.SU_OBS.toPlainText(),self.ui.SU_FECHA.text(),self.ui.SU_TEL.text(),int(self.TextoComboBox(self.ui.SU_TR).split(":")[0]))
        self.DBUnidades()
    def DBEliminar(self):
        if self.indicador=="PR":
            db.borrarPROBLEMAS_RED(int(self.ID))
            self.DBproblemasRed()
            self.DBcerrarMenuAct()
        elif self.indicador=="R":
            db.borrarREPORTES(int(self.ID))
            self.DBreportes()
            self.DBcerrarMenuAct()
        elif self.indicador=="I":
            db.borrarDISPOSITIVOS(int(self.ID))
            self.DBinvernario()
            self.DBcerrarMenuAct()
        elif self.indicador=="P":
            db.borrarPROVEEDOR(int(self.ID))
            self.DBProveedores()
            self.DBcerrarMenuAct()
        elif self.indicador=="RD":
            db.borrarRED(int(self.ID))
            self.DBRedes()
            self.DBcerrarMenuAct()
        elif self.indicador=="CR":
            db.borrarCONTROL_RED_ESPECIFICA(int(self.ID))
            self.DBcerrarMenuAct()
            self.DBControlRed()
        elif self.indicador=="U":
            db.borrarUNIDADES_INTERNET(self.ui.SU_ID.text())
            self.DBcerrarMenuAct()
            self.DBUnidades()
        elif self.indicador=="D":
            db.borrarDEPARTAMENTOS(int(self.ID))
            self.DBcerrarMenuAct()
            self.DBDepartamentos()

    def DBModificar(self):
        if self.indicador=="PR":
            db.modificarPROBLEMAS_RED(int(self.ID),self.ui.PR_DESC.toPlainText(),self.ui.PR_FECHA.text(),self.ui.PR_SOL.toPlainText(),self.TextoComboBox(self.ui.PR_COR),self.ui.PR_FECHAC.text(),int(self.TextoComboBox(self.ui.PR_RED).split(":")[0]),int(self.TextoComboBox(self.ui.PR_DEP).split(":")[0]))
            self.DBproblemasRed()
            self.DBcerrarMenuAct()
        elif self.indicador=="R":
            db.modificarREPORTES(int(self.ID),self.ui.R_Folio.text(),self.ui.R_Descrip.toPlainText(),self.ui.R_Fecha.text(),self.TextoComboBox(self.ui.R_ACUDIO),self.TextoComboBox(self.ui.R_REMOSIT),self.ui.R_MOT.toPlainText(),self.ui.R_TEL.text(),self.ui.R_ATEND.text(),int(self.TextoComboBox(self.ui.R_PROV).split(":")[0]))
            self.DBreportes()
            self.DBcerrarMenuAct()
        elif self.indicador=="I":
            db.modificarDISPOSITIVOS(int(self.ID),self.ui.I_ID.text(),self.ui.I_NOM.text(),self.ui.I_DESC.toPlainText(),self.ui.I_UBI.text(),self.TextoComboBox(self.ui.I_TC),int(self.TextoComboBox(self.ui.I_RED).split(":")[0]),int(self.TextoComboBox(self.ui.I_DEP).split(":")[0]))
            self.DBinvernario()
            self.DBcerrarMenuAct()
        elif self.indicador=="P":
            db.modificarPROVEEDOR(int(self.ID),self.ui.P_N.text(),self.ui.P_DES.toPlainText(),self.ui.P_FECH.text(),self.ui.P_TEL.text(),int(self.TextoComboBox(self.ui.P_TR).split(":")[0]))
            self.DBProveedores()
            self.DBcerrarMenuAct()
        elif self.indicador=="RD":
            db.modificarRED(int(self.ID),self.ui.RD_S.text(),self.ui.RD_PASS.text(),float(self.ui.RD_SUB.text()),float(self.ui.RD_DAJ.text()),int(self.TextoComboBox(self.ui.RD_PROV).split(":")[0]))
            self.DBRedes()
            self.DBcerrarMenuAct()
        elif self.indicador=="CR":
            db.modificarCONTROL_RED_ESPECIFICA(int(self.ID),int(self.TextoComboBox(self.ui.CR_RED).split(":")[0]),self.ui.CR_IP.text(),self.ui.CR_NOINV.text(),self.TextoComboBox(self.ui.CR_STAT),self.ui.CR_MAC.text(),self.ui.CR_RESP.text(),self.ui.CR_FECHA.text(),self.ui.CR_OBS.toPlainText(),self.ui.CR_HOS.text(),self.TextoComboBox(self.ui.CR_VERFI),self.TextoComboBox(self.ui.CR_TIP),int(self.TextoComboBox(self.ui.CR_DEP).split(":")[0]),self.ui.CR_VEND.text())
            self.DBcerrarMenuAct()
            self.DBControlRed()
        elif self.indicador=="U":
            db.modificarUNIDADES_INTERNET(self.ui.SU_ID.text(),self.ui.SU_NOM.text(),self.ui.SU_CORD.text(),self.TextoComboBox(self.ui.SU_SERV),self.ui.SU_PROVE.text(),self.ui.SU_OBS.toPlainText(),self.ui.SU_FECHA.text(),self.ui.SU_TEL.text(),int(self.TextoComboBox(self.ui.SU_TR).split(":")[0]))
            self.DBcerrarMenuAct()
            self.DBUnidades()
        elif self.indicador=="D":
            db.modificarDEPARTAMENTOS(int(self.ID),self.ui.D_NOM.text(),self.ui.D_DES.toPlainText())
            self.DBcerrarMenuAct()
            self.DBDepartamentos()
        self.DBcerrarMenuAct()
        formato="border-radius:10px;padding: 8px 8px;"
        self.ui.GModificarBtn.setStyleSheet(formato)
        self.ui.GEliminarBtn.setStyleSheet(formato)
        self.ui.GModificarBtn.setEnabled(False)
        self.ui.GEliminarBtn.setEnabled(False)
        formato="border-radius:10px;background-color: #20945e;	padding: 8px 8px;"
        self.ui.CR_ADD.setStyleSheet(formato)
        self.ui.D_ADD.setStyleSheet(formato)
        self.ui.I_ADD.setStyleSheet(formato)
        self.ui.P_ADD.setStyleSheet(formato)
        self.ui.G_AgregarBtn.setStyleSheet(formato)
        self.ui.R_ADD.setStyleSheet(formato)
        self.ui.RD_ADD.setStyleSheet(formato)
        self.ui.SU_ADD.setStyleSheet(formato)
        self.ui.D_ADD.setEnabled(True)
        self.ui.I_ADD.setEnabled(True)
        self.ui.R_ADD.setEnabled(True)
        self.ui.P_ADD.setEnabled(True)
        self.ui.CR_ADD.setEnabled(True)
        self.ui.RD_ADD.setEnabled(True)
        self.ui.SU_ADD.setEnabled(True)
        self.ui.G_AgregarBtn.setEnabled(True)
        self.ui.SU_ID.setEnabled(True)
    def TextoComboBox(self,item):
        aux= item.itemText(item.currentIndex())
        return aux 
    ID=None
    def dobleGestion(self):
        self.ID=None
        self.indicadorM=True
        self.DBAbrirMenuAct()
        formato="border-radius:10px;background-color: #20945e;	padding: 8px 8px;"
        self.ui.GModificarBtn.setEnabled(True)
        self.ui.GEliminarBtn.setEnabled(True)
        self.ui.GModificarBtn.setStyleSheet(formato)
        self.ui.GEliminarBtn.setStyleSheet(formato)
        self.ui.D_ADD.setEnabled(False)
        self.ui.I_ADD.setEnabled(False)
        self.ui.R_ADD.setEnabled(False)
        self.ui.P_ADD.setEnabled(False)
        self.ui.CR_ADD.setEnabled(False)
        self.ui.RD_ADD.setEnabled(False)
        self.ui.SU_ADD.setEnabled(False)
        self.ui.SU_ID.setEnabled(False)
        formato="border-radius:10px;padding: 8px 8px;"
        self.ui.CR_ADD.setStyleSheet(formato)
        self.ui.D_ADD.setStyleSheet(formato)
        self.ui.I_ADD.setStyleSheet(formato)
        self.ui.P_ADD.setStyleSheet(formato)
        self.ui.G_AgregarBtn.setStyleSheet(formato)
        self.ui.R_ADD.setStyleSheet(formato)
        self.ui.RD_ADD.setStyleSheet(formato)
        self.ui.SU_ADD.setStyleSheet(formato)

        self.ui.G_AgregarBtn.setEnabled(False)
        row_number=0
        for idx in self.ui.SalidaGestion.selectionModel().selectedIndexes():
            row_number = idx.row()
        aux=[]
        for i in range(0,self.ui.SalidaGestion.columnCount()):
            aux.append(self.ui.SalidaGestion.item(row_number,i).text())
        self.ID=str(aux[0])
        
        if self.indicador=="PR":
            self.ui.PR_DESC.setText(aux[1])
            self.ui.PR_FECHA.setText(aux[2])
            self.ui.PR_SOL.setText(aux[3])
            self.ui.PR_COR.clear()
            self.ui.PR_COR.addItem(aux[4])
            if aux[4]=="No corregido": self.ui.PR_COR.addItem("Corregido")
            elif aux[4]=="Corregido": self.ui.PR_COR.addItem("No corregido")
            else: 
                self.ui.PR_COR.addItem("Corregido")
                self.ui.PR_COR.addItem("No corregido")
            self.ui.PR_FECHAC.setText(aux[5])
            self.comboDB(self.ui.PR_RED,f"SELECT ID_RED FROM PROBLEMAS_RED WHERE ID_PROBLEMA={aux[0]}",db.consultarRED(),aux[6])
            self.comboDB(self.ui.PR_DEP,f"SELECT ID_DEPARTAMENTO FROM PROBLEMAS_RED WHERE ID_PROBLEMA={aux[0]}",db.consultarDEPARTAMENTOS(),aux[7])
            
        elif self.indicador=="R":
            self.ui.R_Folio.setText(aux[1])
            self.ui.R_Descrip.setText(aux[2])
            self.ui.R_Fecha.setText(aux[3])
            self.ui.R_ACUDIO.clear()
            self.ui.R_ACUDIO.addItem(aux[4])
            if aux[4]=="No acudio al sitio": self.ui.R_ACUDIO.addItem("Acudio al sitio")
            elif aux[4]=="Acudio al sitio": self.ui.R_ACUDIO.addItem("No acudio al sitio")
            else: 
                self.ui.R_ACUDIO.addItem("Acudio al sitio")
                self.ui.R_ACUDIO.addItem("No acudio al sitio")
            self.ui.R_REMOSIT.clear()
            self.ui.R_REMOSIT.addItem(aux[5])
            if aux[5]=="Procedimiento de manera Remota": self.ui.R_REMOSIT.addItem("Procedimientos en Sitio")
            elif aux[5]=="Procedimientos en Sitio": self.ui.R_REMOSIT.addItem("Procedimiento de manera Remota")
            else: 
                self.ui.R_REMOSIT.addItem("Procedimiento de manera Remota")
                self.ui.R_REMOSIT.addItem("Procedimientos en Sitio")
            self.ui.R_MOT.setText(aux[6])
            self.ui.R_TEL.setText(aux[7])
            self.ui.R_ATEND.setText(aux[8])
            self.comboDB(self.ui.R_PROV,f"SELECT ID_PROVEEDOR FROM REPORTES WHERE ID_REPORTE={aux[0]}",db.consultarPROVEEDOR(),aux[9])
        elif self.indicador=="I":
            self.ui.I_ID.setText(aux[1])
            self.ui.I_NOM.setText(aux[2])
            self.ui.I_DESC.setText(aux[3])
            self.ui.I_UBI.setText(aux[4])
            self.ui.I_TC.clear()
            self.ui.I_TC.addItem(aux[5])
            if aux[5]=="Inalámbrica": 
                self.ui.I_TC.addItem("Alambrica")
                self.ui.I_TC.addItem("No Aplica")
            elif aux[5]=="Alambrica": 
                self.ui.I_TC.addItem("Inalámbrica")
                self.ui.I_TC.addItem("No Aplica")
            elif aux[5]=="No aplica": 
                self.ui.I_TC.addItem("Alambrica")
                self.ui.I_TC.addItem("Inalámbrica")
            else: 
                self.ui.I_TC.addItem("Alambrica")
                self.ui.I_TC.addItem("Inalámbrica")
                self.ui.I_TC.addItem("No Aplica")
            self.comboDB(self.ui.I_RED,f"SELECT ID_RED FROM DISPOSITIVOS WHERE ID_DISPOSITIVO={aux[0]}",db.consultarRED(),aux[6])
            self.comboDB(self.ui.I_DEP,f"SELECT ID_DEPARTAMENTO FROM DISPOSITIVOS WHERE ID_DISPOSITIVO={aux[0]}",db.consultarDEPARTAMENTOS(),aux[7])
        elif self.indicador=="P":
            self.ui.P_N.setText(aux[1])
            self.ui.P_DES.setText(aux[2])
            self.ui.P_FECH.setText(aux[3])
            self.ui.P_TEL.setText(aux[4])
            self.comboDB(self.ui.P_TR,f"SELECT ID_TRANSMISION FROM PROVEEDOR WHERE ID_PROVEEDOR={aux[0]}",db.consultarTRANSMISION(),aux[5])
        elif self.indicador=="RD":
            self.ui.RD_S.setText(aux[1])
            self.ui.RD_PASS.setText(aux[2])
            self.ui.RD_SUB.setText(aux[3])
            self.ui.RD_DAJ.setText(aux[4])
            self.comboDB(self.ui.RD_PROV,f"SELECT ID_PROVEEDOR FROM RED WHERE ID_RED={aux[0]}",db.consultarPROVEEDOR(),aux[5])
        elif self.indicador=="CR":
            self.ui.CR_FECHA.setText(aux[1])
            self.comboDB(self.ui.CR_RED,f"SELECT ID_RED FROM CONTROL_RED_ESPECIFICA WHERE ID_CONTROL_RED_ESPECIFICA={aux[0]}",db.consultarRED(),aux[2])
            self.ui.CR_NOINV.setText(aux[3])
            self.ui.CR_STAT.clear()
            self.ui.CR_STAT.addItem(aux[4])
            if aux[4]=="Libre": self.ui.CR_STAT.addItem("Ocupado")
            elif aux[4]=="Ocupado": self.ui.CR_STAT.addItem("Libre")
            else: 
                self.ui.CR_STAT.addItem("Libre")
                self.ui.CR_STAT.addItem("Ocupado")
            self.ui.CR_HOS.setText(aux[5])
            self.ui.CR_IP.setText(aux[6])
            self.ui.CR_VEND.setText(aux[7])
            self.ui.CR_MAC.setText(aux[8])
            self.ui.CR_RESP.setText(aux[9])
            self.ui.CR_OBS.setText(aux[10])
            self.ui.CR_VERFI.clear()
            self.ui.CR_VERFI.addItem(aux[11])
            if aux[11]=="No verificado": 
                self.ui.CR_VERFI.addItem("No permitido")
                self.ui.CR_VERFI.addItem("Verificado")
            elif aux[11]=="Verificado": 
                self.ui.CR_VERFI.addItem("No verificado")
                self.ui.CR_VERFI.addItem("No permitido")
            elif aux[11]=="No permitido": 
                self.ui.CR_VERFI.addItem("Verificado")
                self.ui.CR_VERFI.addItem("No verificado")
            else: 
                self.ui.CR_VERFI.addItem("No verificado")
                self.ui.CR_VERFI.addItem("Verificado")
                self.ui.CR_VERFI.addItem("No permitido")
            self.ui.CR_TIP.clear()
            self.ui.CR_TIP.addItem(aux[12])
            if aux[12]=="Móvil": 
                self.ui.CR_TIP.addItem("Laptop")
                self.ui.CR_TIP.addItem("Escritorio")
            elif aux[12]=="Laptop": 
                self.ui.CR_TIP.addItem("Móvil")
                self.ui.CR_TIP.addItem("Escritorio")
            elif aux[12]=="Escritorio": 
                self.ui.CR_TIP.addItem("Móvil")
                self.ui.CR_TIP.addItem("Laptop")
            else: 
                self.ui.CR_TIP.addItem("Móvil")
                self.ui.CR_TIP.addItem("Laptop")
                self.ui.CR_TIP.addItem("Escritorio")
            self.comboDB(self.ui.CR_DEP,f"SELECT ID_DEPARTAMENTO FROM CONTROL_RED_ESPECIFICA WHERE ID_CONTROL_RED_ESPECIFICA={aux[0]}",db.consultarDEPARTAMENTOS(),aux[13])

        elif self.indicador=="U":
            self.ui.SU_ID.setText(aux[0])
            self.ui.SU_NOM.setText(aux[1])
            self.ui.SU_CORD.setText(aux[2])
            self.ui.SU_SERV.clear()
            self.ui.SU_SERV.addItem(aux[3])
            if aux[3]=="Sin Servicio": self.ui.SU_SERV.addItem("Con Servicio")
            elif aux[3]=="Con Servicio": self.ui.SU_SERV.addItem("Sin Servicio")
            else: 
                self.ui.SU_SERV.addItem("Con Servicio")
                self.ui.SU_SERV.addItem("Sin Servicio")
            self.ui.SU_PROVE.setText(aux[4])
            self.ui.SU_OBS.setText(aux[5])
            self.ui.SU_FECHA.setText(aux[6])
            self.ui.SU_TEL.setText(aux[7])
            self.comboDB(self.ui.SU_TR,f"SELECT ID_TRANSMISION FROM UNIDADES_INTERNET WHERE ID_UNIDAD='{aux[0]}'",db.consultarTRANSMISION(),aux[8])
        elif self.indicador=="D":
            self.ui.D_NOM.setText(aux[1])
            self.ui.D_DES.setText(aux[2])
    def limpiarCampos(self):
        if self.indicadorV=="PR":
            self.ui.PR_DESC.setText("")
            self.ui.PR_FECHA.setText("")
            self.ui.PR_SOL.setText("")
            self.ui.PR_DEP.clear()
            self.ui.PR_COR.clear()
            self.ui.PR_COR.addItem("Corregido")
            self.ui.PR_COR.addItem("No corregido")
            self.ui.PR_FECHAC.setText("")

            self.comboDBD(self.ui.PR_RED,db.consultarRED())
            self.comboDBD(self.ui.PR_DEP,db.consultarDEPARTAMENTOS())
            
        elif self.indicador=="R":
            self.ui.R_Folio.setText("")
            self.ui.R_Descrip.setText("")
            self.ui.R_Fecha.setText("")
            self.ui.R_ACUDIO.clear()
            self.ui.R_ACUDIO.addItem("Acudio al sitio")
            self.ui.R_ACUDIO.addItem("No acudio al sitio")
            self.ui.R_REMOSIT.clear()
            self.ui.R_REMOSIT.addItem("Procedimiento de manera Remota")
            self.ui.R_REMOSIT.addItem("Procedimientos en Sitio")
            self.ui.R_MOT.setText("")
            self.ui.R_TEL.setText("")
            self.ui.R_ATEND.setText("")
            self.comboDBD(self.ui.R_PROV,db.consultarPROVEEDOR())
        elif self.indicador=="I":
            self.ui.I_ID.setText("")
            self.ui.I_NOM.setText("")
            self.ui.I_DESC.setText("")
            self.ui.I_UBI.setText("")
            self.ui.I_TC.clear()
            self.ui.I_TC.addItem("Alambrica")
            self.ui.I_TC.addItem("Inalámbrica")
            self.ui.I_TC.addItem("No Aplica")
            self.comboDBD(self.ui.I_RED,db.consultarRED())
            self.comboDBD(self.ui.I_DEP,db.consultarDEPARTAMENTOS())
        elif self.indicador=="P":
            self.ui.P_N.setText("")
            self.ui.P_DES.setText("")
            self.ui.P_FECH.setText("")
            self.ui.P_TEL.setText("")
            self.comboDBD(self.ui.P_TR,db.consultarTRANSMISION())
        elif self.indicador=="RD":
            self.ui.RD_S.setText("")
            self.ui.RD_PASS.setText("")
            self.ui.RD_SUB.setText("")
            self.ui.RD_DAJ.setText("")
            self.comboDBD(self.ui.RD_PROV,db.consultarPROVEEDOR())
        elif self.indicador=="CR":
            self.ui.CR_FECHA.setText("")
            self.comboDBD(self.ui.CR_RED,db.consultarCONTROL_RED_ESPECIFICA())
            self.ui.CR_NOINV.setText("")
            self.ui.CR_STAT.clear()
            self.ui.CR_VERFI.addItem("No verificado")
            self.ui.CR_VERFI.addItem("Verificado")
            self.ui.CR_VERFI.addItem("No permitido")
            self.ui.CR_HOS.setText("")
            self.ui.CR_IP.setText("")
            self.ui.CR_VEND.setText("")
            self.ui.CR_MAC.setText("")
            self.ui.CR_RESP.setText("")
            self.ui.CR_OBS.setText("")
            self.ui.CR_TIP.clear()
            self.ui.CR_TIP.addItem("Móvil")
            self.ui.CR_TIP.addItem("Laptop")
            self.ui.CR_TIP.addItem("Escritorio")

        elif self.indicador=="U":
            self.ui.SU_ID.setText("")
            self.ui.SU_NOM.setText("")
            self.ui.SU_CORD.setText("")
            self.ui.SU_SERV.clear()
            self.ui.SU_SERV.addItem("Con Servicio")
            self.ui.SU_SERV.addItem("Sin Servicio")
            self.ui.SU_PROVE.setText("")
            self.ui.SU_OBS.setText("")
            self.ui.SU_FECHA.setText("")
            self.ui.SU_TEL.setText("")
            self.comboDBD(self.ui.SU_TR,db.consultarTRANSMISION())
        elif self.indicador=="D":
            self.ui.D_NOM.setText("")
            self.ui.D_DES.setText("")
    def comboDB(self,combo,consulta,dep,val):
        combo.clear()
        a=db.consultar(consulta)
        combo.addItem(str(a[0][0])+":"+val)
        for i in dep: 
                if str(a[0][0])+":"+val!=str(i[0])+":"+i[1]:
                    combo.addItem(str(i[0])+":"+i[1])
    def comboDBD(self,combo,dep):
        combo.clear()
        for i in dep: combo.addItem(str(i[0])+":"+i[1])
    def Filtrar(self):
        self.Filtrado=self.ui.GFiltrado.text()
        if self.Filtrado=="" or self.Filtrado.replace(" ","").replace("\t","")=="":
            self.Filtrado="*_*_*"
        if self.indicador=="PR":  self.DBproblemasRed()
        elif self.indicador=="R": self.DBreportes()
        elif self.indicador=="I": self.DBinvernario()
        elif self.indicador=="P": self.DBProveedores()
        elif self.indicador=="RD": self.DBRedes()
        elif self.indicador=="C": self.DBConexiones()
        elif self.indicador=="H": self.DBHosts()
        elif self.indicador=="CR": self.DBControlRed()
        elif self.indicador=="U": self.DBUnidades()
        elif self.indicador=="D": self.DBDepartamentos()
    def redesR(self):
        redes= sb.run(["netsh", "wlan", "show","profile"], capture_output=True, text=True).stdout
        ls=redes.split("\n")
        redes=[]
        for i in ls:
            if ":" in i  and "actual" not in i and "current" not in i:
                aux=i.split(":")[1]
                if aux!="":
                    redes.append(aux.strip())
        return redes
    def CambiosRed(self):
        while self.inf.SALIR!=True:
            SSID=rs.obtener_SSID(self.inf.INTERFAZ)
            INTER=rs.Obtener_Dispositivos(False)
            auxIP=""
            for dis in INTER:
                if dis[0]==self.inf.INTERFAZ:
                    auxIP=dis[3]
                    self.inf.auxGate=dis[5] 
                    break
            if SSID=="SIN CONEXION" or (self.inf.auxGate=="--" or self.inf.auxGate==""):
                print("[i] No hay conexión en la interfaz "+self.inf.INTERFAZ)
                self.inf.SSID="SIN CONEXION"
                self.inf.CONEXION=False
                self.inf.IPV4="--"
                self.inf.GATEWAY="--"
                self.ui.UserBtn.setText("")
                self.ui.label_5.setText(f"<p align='right'><b>Red:</b> {self.inf.SSID}&nbsp; &nbsp; &nbsp; <b>IPv4:</b> {self.inf.IPV4}&nbsp; &nbsp; &nbsp; <b>Gateway:</b> {self.inf.GATEWAY}</p>")
                self.actualizarDatosDash()
                self.inf.CAMBIOS=False
                self.ui.SpeedRed.setText(self.inf.SSID)
            else:
                if SSID=='NO TIENE' and "ETHER" in self.inf.INTERFAZ.upper():
                    if self.inf.auxGate != self.inf.GATEWAY:
                        print("[!] Cambios en red. (Cableada)")
                        self.inf.CAMBIOS=True
                        self.inf.IPV4=auxIP
                        self.inf.GATEWAY=self.inf.auxGate     
                        try:
                            SSID=str(socket.getfqdn(self.inf.GATEWAY))
                            if SSID==self.inf.GATEWAY:
                                try:
                                    SSID=str(rs.get_mac_details(rs.escanearARP_U(self.inf.GATEWAY,self.inf.INTERFAZ,3,False)[1]))
                                except:
                                    SSID=self.inf.GATEWAY
                        except:
                            try:
                                SSID=str(rs.get_mac_details(rs.escanearARP_U(self.inf.GATEWAY,self.inf.INTERFAZ,3,False)[1]))
                            except:
                                SSID=self.inf.GATEWAY        
                        self.inf.SSID=SSID
                        self.inf.CONEXION=True
                        aux=db.consultarRED()
                        encontrado=False
                        for i in aux:
                            if i[1]==self.inf.SSID:
                                encontrado=True
                        if encontrado==False and self.inf.SSID!='POR DEFECTO' and self.inf.SSID!='SIN CONEXION':
                            db.insertarRED(self.inf.SSID,"Pendiente",0,0,1)
                        self.ui.label_5.setText(f"<p align='right'><b>Red:</b> {self.inf.SSID}&nbsp; &nbsp; &nbsp; <b>IPv4:</b> {self.inf.IPV4}&nbsp; &nbsp; &nbsp; <b>Gateway:</b> {self.inf.GATEWAY}</p>")
                        self.actualizarDatosDash()
                        self.ui.UserBtn.setText("")
                        self.inf.CAMBIOS=False
                        self.ui.SpeedRed.setText(self.inf.SSID)
                        self.tab.emit(1)
                elif self.inf.SSID!=SSID and SSID != "NO TIENE":
                    print("[!] Cambios en red. (Inalambrica)")
                    self.inf.CAMBIOS=True
                    self.inf.IPV4=auxIP
                    self.inf.GATEWAY=self.inf.auxGate      
                    self.inf.SSID=SSID
                    self.inf.CONEXION=True
                    self.ui.UserBtn.setText("")
                    sleep(1)
                    aux=db.consultarRED()
                    encontrado=False
                    for i in aux:
                        if i[1]==self.inf.SSID:
                            encontrado=True
                    if encontrado==False and self.inf.SSID!='POR DEFECTO' and self.inf.SSID in self.redesR():
                        db.insertarRED(self.inf.SSID,"Pendiente",0,0,1)
                    self.ui.label_5.setText(f"<p align='right'><b>Red:</b> {self.inf.SSID}&nbsp; &nbsp; &nbsp; <b>IPv4:</b> {self.inf.IPV4}&nbsp; &nbsp; &nbsp; <b>Gateway:</b> {self.inf.GATEWAY}</p>")
                    self.actualizarDatosDash()
                    self.inf.CAMBIOS=False
                    self.ui.SpeedRed.setText(self.inf.SSID)
                    self.tab.emit(1)
            sleep(1)
    def RGeneral(self):
        archivo="Gtmp.xlsx"
        wb = Workbook()
        hoja=wb.active
        hoja.title="Redes"
        hoja.append(('ID','SSID','CONTRASEÑA','SUBIDA','BAJADA','PROVEEDOR'))
        consulta=db.consultarRED()
        for i in consulta:
            hoja.append(i)
        hoja1=wb.create_sheet("Pruebas")
        hoja1.append(('ID','FECHA','HORA','HOST','SPONSOR','RED','SUBIDA','BAJADA','LATENCIA(S)','LATENCIA(B)','PING','RESULTADO'))
        consulta=db.consultarPRUEBAS_VELOCIDAD()
        for i in consulta:
            hoja1.append(i)
        hoja2=wb.create_sheet("Reportes")
        hoja2.append(('ID','FOLIO','DESCRIPCION','FECHA','ACUDIO','REMOTO/SITIO','MOTIVO','TELEFONO','ATENDIO','PROVEEDOR'))
        consulta=db.consultarREPORTES()
        for i in consulta:
            hoja2.append(i)
        hoja3=wb.create_sheet("Unidades Internet")
        hoja3.append(('ID','NOMBRE','COORDINACION','CON INTENERT','PROVEEDOR','OBSERVACIONES','FECHA','TELEFONO','TIPO TRANSMISIÓN'))
        consulta=db.consultarUNIDADES_INTERNET()
        for i in consulta:
            hoja3.append(i)
        hoja4=wb.create_sheet("Control de red")
        hoja4.append(('ID','FECHA','SSID','NO. INVENTARIO','ESTADO','HOST','IPV4','VENDOR','MAC','USUARIO','OBSERVACIONES','VERIFICADO','TIPO','DEPARTAMENTO'))
        consulta=db.consultarCONTROL_RED_ESPECIFICA()
        for i in consulta:
            hoja4.append(i)
        hoja5=wb.create_sheet("Numero de conexiones")
        hoja5.append(('ID','SSID','FECHA','CONEXIONES'))
        consulta=db.consultarHISTORIAL_NUM_CONEXIONES()
        for i in consulta:
            hoja5.append(i)
        hoja6=wb.create_sheet("Historial de dispositivos")
        hoja6.append(('ID','FECHA','SSID','HOST','IPV4','MAC','VENDOR'))
        consulta=db.consultarHISTORIAL_CONEXIONES()
        for i in consulta:
            hoja6.append(i)
        hoja7=wb.create_sheet("Proveedores")
        hoja7.append(('ID','NOMBRE','DESCRIPCION','FECHA','TELEFONO','TIPO'))
        consulta=db.consultarPROVEEDOR()
        for i in consulta:
            hoja7.append(i)
        hoja8=wb.create_sheet("Inventario")
        hoja8.append(('ID','NO.INVENTARIO','NOMBRE','DESCRIPCION','UBICACION','TIPO','RED','DEPARTAMENTO'))
        consulta=db.consultarDISPOSITIVOS()
        for i in consulta:
            hoja8.append(i)
        hoja9=wb.create_sheet("Departamentos")
        hoja9.append(('ID','NOMBRE','DESCRIPCION'))
        consulta=db.consultarDEPARTAMENTOS()
        for i in consulta:
            hoja9.append(i)
        hoja10=wb.create_sheet("Problemas red")
        hoja10.append(('ID','Descripción','Fecha','Solución','Corregido','Fecha corrección','Red','Departamento'))
        consulta=db.consultarPROBLEMAS_RED()
        for i in consulta:
            hoja10.append(i)
        wb.save(archivo)
        os.system("start "+archivo)
    def RVelocidad(self):
        archivo="Vtmp.xlsx"
        wb = Workbook()
        hoja=wb.active
        hoja.title="Pruebas"
        hoja.append(('ID','FECHA','HORA','HOST','SPONSOR','RED','SUBIDA','BAJADA','LATENCIA(S)','LATENCIA(B)','PING','RESULTADO'))
        consulta=db.consultarPRUEBAS_VELOCIDAD()
        for i in consulta:
            hoja.append(i)
        wb.save(archivo)
        os.system("start "+archivo)
    def RServicio(self):
        archivo="Stmp.xlsx"
        wb = Workbook()
        hoja=wb.active
        hoja.title="Reportes"
        hoja.append(('ID','FOLIO','DESCRIPCION','FECHA','ACUDIO','REMOTO/SITIO','MOTIVO','TELEFONO','ATENDIO','PROVEEDOR'))
        consulta=db.consultarREPORTES()
        for i in consulta:
            hoja.append(i)
        wb.save(archivo)
        os.system("start "+archivo)
    def RClues(self):
        archivo="Ctmp.xlsx"
        wb = Workbook()
        hoja=wb.active
        hoja.title="Unidades con Internet"
        hoja.append(('ID','NOMBRE','CORD','CON INTENERT','PROVEEDOR','OBSERVACIONES','FECHA','TELEFONO','TIPO TRANSMISIÓN'))
        consulta=db.consultarUNIDADES_INTERNET()
        for i in consulta:
            hoja.append(i)
        wb.save(archivo)
        os.system("start "+archivo)
    def RDisp(self):
        archivo="CRtmp.xlsx"
        wb = Workbook()
        hoja=wb.active
        hoja.title="HOST"
        hoja.append(('ID','FECHA','SSID','NO. INVENTARIO','ESTADO','HOST','IPV4','VENDOR','MAC','USUARIO','OBSERVACIONES','VERIFICADO','TIPO','DEPARTAMENTO'))
        consulta=db.consultarCONTROL_RED_ESPECIFICA()
        for i in consulta:
            hoja.append(i)
        wb.save(archivo)
        os.system("start "+archivo)
    def RHistorial(self):
        archivo="Htmp.xlsx"
        wb = Workbook()
        hoja1=wb.active
        hoja1.title="Numero de conexiones"
        hoja1.append(('ID','SSID','FECHA','CONEXIONES'))
        consulta=db.consultarHISTORIAL_NUM_CONEXIONES()
        for i in consulta:
            hoja1.append(i)
        hoja2=wb.create_sheet("Historial de dispositivos")
        hoja2.append(('ID','FECHA','SSID','HOST','IPV4','MAC','VENDOR'))
        consulta=db.consultarHISTORIAL_CONEXIONES()
        for i in consulta:
            hoja2.append(i)
        wb.save(archivo)
        os.system("start "+archivo)
    def RProveedores(self):
        archivo="Ptmp.xlsx"
        wb = Workbook()
        hoja=wb.active
        hoja.title="Proveedores"
        hoja.append(('ID','NOMBRE','DESCRIPCION','FECHA','TELEFONO','TIPO'))
        consulta=db.consultarPROVEEDOR()
        for i in consulta:
            hoja.append(i)
        wb.save(archivo)
        os.system("start "+archivo)
    def RInventario(self):
        archivo="Itmp.xlsx"
        wb = Workbook()
        hoja=wb.active
        hoja.title="Inventario"
        hoja.append(('ID','NO.INVENTARIO','NOMBRE','DESCRIPCION','UBICACION','TIPO','RED','DEPARTAMENTO'))
        consulta=db.consultarDISPOSITIVOS()
        for i in consulta:
            hoja.append(i)
        wb.save(archivo)
        os.system("start "+archivo)
    def RProblemas(self):
        archivo="PRtmp.xlsx"
        wb = Workbook()
        hoja=wb.active
        hoja.title="Problemas"
        hoja.append(('ID','Descripción','Fecha','Solución','Corregido','Fecha corrección','Red','Departamento'))
        consulta=db.consultarPROBLEMAS_RED()
        for i in consulta:
            hoja.append(i)
        wb.save(archivo)
        os.system("start "+archivo)
    def BRHistorial(self):
        db.ejecutarAccion("DELETE FROM HISTORIAL_NUM_CONEXIONES")
        db.ejecutarAccion("DELETE FROM HISTORIAL_CONEXIONES")
        self.inf.Errores.append([f"{dt.now().hour}:{dt.now().minute}","Base de Datos","Se ha borrado el historial de red."])
        self.mostrarErrores()
    def BRServicio(self):
        db.ejecutarAccion("DELETE FROM REPORTES")
        self.inf.Errores.append([f"{dt.now().hour}:{dt.now().minute}","Base de Datos","Se ha borrado los reportes de servicio."])
    def BRPruebas(self):
        db.ejecutarAccion("DELETE FROM PRUEBAS_VELOCIDAD")
        self.inf.Errores.append([f"{dt.now().hour}:{dt.now().minute}","Base de Datos","Se ha borrado las pruebas de red."])
    def BRProblemas(self):
        db.ejecutarAccion("DELETE FROM PROBLEMAS_RED")
        self.inf.Errores.append([f"{dt.now().hour}:{dt.now().minute}","Base de Datos","Se ha borrado los problemas de red."])
    def VMac(self):
        redes= sb.run(["getmac"], capture_output=True, text=True).stdout
        self.ui.SalidaVar.setHtml(str(redes.replace("\n","<br>")))
    def VARP(self):
        A,disp,ips=rs.escanearARP(self.inf.GATEWAY+"/24",self.inf.INTERFAZ)
        ARP=""
        ARP+="Respuesta a solicitudes ARP<br>"
        for i in disp:
            ARP+=i[0]+"&nbsp;&nbsp;&nbsp;"+i[1]+"<br>"
        self.ui.SalidaVar.setHtml(ARP)
    def speedtest(self):
        WebOpen('https://www.speedtest.net/')
    def fast(self):
        WebOpen('https://fast.com/es/')
    def MRed(self):
        os.system("start ncpa.cpl")
    def MTareas(self):
        os.system("start taskmgr")
    def RCompartidos(self):
        WebOpen(self.inf.GATEWAY)
    def cmd(self):
        os.system("start cmd")
    def errores(self):
        Widget=QWidget()
        aux=QVBoxLayout()
        if len(self.inf.Errores)>0:
            for i in self.inf.Errores:
                auxWid=QFrame()
                auxW=QGridLayout()
                label=QLabel()
                pixmap = QPixmap('Imagenes/ALERT.png').scaled(50,50)
                label.setPixmap(pixmap)
                auxW.addWidget(label,0,0)
                auxW.addWidget(QLabel(i[0]),0,1)
                auxW.addWidget(QLabel(i[1]),0,2)
                auxW.addWidget(QLabel(i[2]),0,3)
                auxWid.setLayout(auxW)
                aux.addWidget(auxWid)
        else:
            auxWid=QFrame()
            auxW=QGridLayout()
            label=QLabel()
            pixmap = QPixmap('Imagenes/CORRECTO.png').scaled(50,50)
            label.setPixmap(pixmap)
            auxW.addWidget(label,0,0)
            auxW.addWidget(QLabel("No hay nada nuevo que informar."),0,1)
            auxWid.setLayout(auxW)
            aux.addWidget(auxWid)
        Widget.setLayout(aux)
        self.ui.NotificacionesW.setWidget(Widget)
    def SignalAlerta(self):
        self.Alerta()
    def mostrarErrores(self):
        if len(self.inf.Errores)>0:
                self.SignalA.emit(1)
    def BorrarNo(self):
        self.inf.Errores=[]
        Widget=QWidget()
        aux=QVBoxLayout()
        auxWid=QFrame()
        auxW=QGridLayout()
        label=QLabel()
        pixmap = QPixmap('Imagenes/CORRECTO.png').scaled(50,50)
        label.setPixmap(pixmap)
        auxW.addWidget(label,0,0)
        auxW.addWidget(QLabel("No hay nada nuevo que informar."),0,1)
        auxWid.setLayout(auxW)
        aux.addWidget(auxWid)
        Widget.setLayout(aux)
        self.ui.NotificacionesW.setWidget(Widget)
        Widget.setLayout(aux)
        self.ui.NotificacionesW.setWidget(Widget)
        
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()

    sys.exit(app.exec())